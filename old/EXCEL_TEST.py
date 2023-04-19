import tkinter as tk
from tkinter import filedialog
import openpyxl
from datetime import datetime, timedelta


# 불러오기 창 생성
root = tk.Tk()
root.withdraw()
file_path = filedialog.askopenfilename()

# 선택된 파일 열기
org_workbook = openpyxl.load_workbook(file_path)
org_worsheet = org_workbook.active

# B 엑셀 파일 열기
save_file_path = 'C:\\excel\\'

# B엑셀 파일을 열어서 2행부터 데이터 입력
save_workbook = openpyxl.load_workbook(save_file_path+"제주귤팜양식.xlsx")
sav_worksheet = save_workbook.active
test_int_1 = int(3)
test_int_2 = int(0)
temp12 = ''
temp13 = ''

# A 엑셀의 3번째 행부터 B 엑셀 2번째 행부터 데이터 복사
for row in range(3, org_worsheet.max_row + 1):
    print("1번째 for test_int_1", test_int_1)
    print("1번째 for test_int_2", test_int_2)
    for col in range(1, org_worsheet.max_column + 1):

        if col == 9:
            sav_worksheet.cell(row=row-1, column=1).value = org_worsheet.cell(row, column=col).value
        
        if col == 10:
            sav_worksheet.cell(row=row-1, column=2).value = org_worsheet.cell(row, column=col).value    
        
        if col == 12:
            temp12 = org_worsheet.cell(row, column=col).value
        
        if col == 13:
            temp13 = org_worsheet.cell(row, column=col).value
            temp14 = str(temp12) + str(temp13)
            sav_worksheet.cell(row=row-1, column=6).value = temp14
        
        if col == 47:
            sav_worksheet.cell(row=row-1, column=5).value = org_worsheet.cell(row, column=col).value
      
    test_int_1 = test_int_1 +1

now_time = datetime.now()
save_fin_name = f"제주귤팜양식_{now_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
save_fin_file_path = save_file_path+save_fin_name
save_workbook.save(save_fin_file_path)
# Bworkbook.save("B.xlsx")


'''
import openpyxl
from datetime import datetime, timedelta
from tkinter import filedialog, Tk

root = Tk()
root.withdraw()
org_file_path = filedialog.askopenfilename()

org_workbook = openpyxl.load_workbook(org_file_path)
org_worsheet = org_workbook['발주발송관리']
# worksheet = org_workbook.active


save_file_path = 'C:\\Users\\김선민\\Documents\\JJGP\\'
# B엑셀 파일을 열어서 2행부터 데이터 입력
save_workbook = openpyxl.load_workbook(save_file_path+"B.xlsx")
save_worksheet = save_workbook['Sheet1']
# save_worksheet = save_workbook.active
current_row = 2  # 2행부터 시작
org_int = int(3)
save_int = int(2)

for row in org_worsheet.iter_rows(min_row=3, values_only=True):
    # 특정 엑셀의 3행의 데이터를 가져와서 B엑셀에 저장
    print(org_int,"번째 진행")
    get_1 = "I" + str(org_int)      # 받으시는 분
    get_2 = "J" + str(org_int)      # 받는분 전화
    get_3 = "AU" + str(org_int)     # 받는분 우편번호
    get_4_1 = "L" + str(org_int)    # 받는분 주소 1
    get_4_2 = "M" + str(org_int)    # 받는분 주소 2
    
    get_data_1 = org_worsheet[get_1].value
    get_data_2 = org_worsheet[get_2].value
    get_data_3 = org_worsheet[get_3].value
    get_data_4_1 = org_worsheet[get_4_1].value
    get_data_4_2 = org_worsheet[get_4_2].value
    
    get_data_4_1 = str(get_data_4_1)
    get_data_4_2 = str(get_data_4_2)
    
    get_data_4 = get_data_4_1 + get_data_4_2 # 받는 분 주소 최종
    
    input_1 = "A" + str(save_int)   # 받으시는 분
    input_2 = "B" + str(save_int)   # 받는분 전화
    input_3 = "E" + str(save_int)   # 받는분 우편번호
    input_4 = "F" + str(save_int)   # 받는 분 주소 최종
    
    
    
    save_worksheet[input_1].value = get_data_1
    save_worksheet[input_2].value = get_data_2
    save_worksheet[input_3].value = get_data_3
    save_worksheet[input_4].value = get_data_4
    
    org_int = org_int + 1
    save_int = save_int + 1




# deal_out_excel_upload_name_now = datetime.now()
# b_save_file_path =b_file_path

now_time = datetime.now()
save_fin_name = f"제주귤팜양식_{now_time.strftime('%Y%m%d_%H%M%S')}.xlsx"
save_fin_file_path = save_file_path+save_fin_name
save_workbook.save(save_fin_file_path)
# Bworkbook.save("B.xlsx")

'''


'''
# 특정 엑셀 파일 열기
specific_wb = openpyxl.load_workbook('specific.xlsx')
specific_ws = specific_wb.active

# B 엑셀 파일 열기
b_wb = openpyxl.load_workbook('B.xlsx')
b_ws = b_wb.active

# 특정 엑셀에서 데이터 가져오기
for row in specific_ws.iter_rows(min_row=3):
    # B 엑셀에서 데이터 입력하기
    b_ws.append([cell.value for cell in row])
    
    # 데이터 업데이트
    b_row = b_ws.max_row
    b_ws.cell(row=b_row, column=1, value=row[8].value)
    b_ws.cell(row=b_row, column=2, value=row[9].value)
    b_ws.cell(row=b_row, column=5, value=row[48].value)
    b_ws.cell(row=b_row, column=6, value=row[11].value + row[12].value)
    
    # 특정 엑셀의 4행이 있을 경우 B 엑셀의 3행에 데이터 저장
    if specific_ws.cell(row=4, column=1).value:
        b_ws.cell(row=3, column=1, value=specific_ws.cell(row=4, column=1).value)
        b_ws.cell(row=3, column=2, value=specific_ws.cell(row=4, column=2).value)
        b_ws.cell(row=3, column=5, value=specific_ws.cell(row=4, column=5).value)
        b_ws.cell(row=3, column=6, value=specific_ws.cell(row=4, column=11).value + specific_ws.cell(row=4, column=12).value)

# 파일 저장
b_wb.save('B.xlsx')
'''
