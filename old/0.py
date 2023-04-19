from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
import selenium.webdriver.support.ui as ui
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog

from openpyxl.styles import Font

import pyautogui
import time
import sys
import json
import re
import openpyxl
import subprocess
import platform
import os.path
from selenium.webdriver.chrome.service import Service

import gspread
from oauth2client.service_account import ServiceAccountCredentials

import requests
from bs4 import BeautifulSoup



#########################################################################################################################
# 사입 요청 파일 다운로드
# https://docs.google.com/spreadsheets/d/19X1duCg7N2npHQHGu_pPcDaji_9pDWdI/edit#gid=1830395100
# 해당 엑셀 파일을 c:\test\ 에 저장



"""
##################### 중요!!!!!
# 테스트 전 복사해서 선언해야 함 deal_test_saip_excel_upload, buyer_wsIdx_name

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화11.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23163' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화12.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23164' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화13.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23165' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화14.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23166' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화15.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23167' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화16.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23168' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화17.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23169' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화18.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23170' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화19.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23171' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값

"""
# deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
# buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값


# 불러오기 창 생성
root = tk.Tk()
root.withdraw()
# excel_upload_file_path = filedialog.askopenfilename()
deal_test_saip_excel_upload = filedialog.askopenfilename()

# 선택된 파일 열기

# deal_test_saip_excel_upload = openpyxl.load_workbook(excel_upload_file_path)
# org_workbook = openpyxl.load_workbook(excel_upload_file_path)
# org_worsheet = org_workbook.active

# 사입 요청한 도매 wsIdx 값

while True:
    buyer_wsIdx_name = input("도매 wsIndex 숫자를 눌러주세요: ")
    if len(buyer_wsIdx_name) == 5:        
        break    
    else:
        print("5자리 숫자를 입력해주세요.")


while True:
    test_google_log = int(input("빠른 입고 작업을 원하면 1, 상세한 입고 작업(구글 로그 확인)을 원하면 9를 눌려주세요: "))
    if test_google_log == 1:
        print("빠른 입고 작업입니다.")
        break
    elif test_google_log == 9:
        print("상세한 입고 작업(구글 로그 확인)입니다.")
        break
    else:
        print("1 또는 9를 입력해주세요.")



#########################################################################################################################
# 테스트 기본 설정
#########################################################################################################################
# 자동화 실행 시 기본 정보 로드 : ex)계정 정보 입력, 구글 접속 정보
deal_admin_login_id = ''
deal_admin_login_password = ''
deal_admin_url = 'https://dealibird.qa.sinsang.market/ssm_admins/sign_in'
deal_seller_login_id = ''
deal_seller_login_password = ''
deal_seller_url = 'https://vat.qa.sinsang.market/'

# WMS 테스트 기본 설정
wms_login_id = '' 									# WMS 로그인 ID
wms_login_passWord = ''  						# WMS 로그인 비번
wms_url = 'https://matrix-web.qa.sinsang.market/signin'


test_os = platform.system()
if test_os == "Windows":
    print("윈도우입니다.", test_os)

    # 로그인, 구글 연동 정보등 초시 세팅 정보 엑셀 파일
    info_file_path = 'C:\\test\\info.xlsx'

    # 출고 관련 정보(배송 요청) : 원본 파일    
    deal_out_excel_upload_path = r"c:\test\auto_out_org_file.xlsx" # Excel 파일 경로 설정

    # 출고에 전달할 정보 : 원본 파일
    deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정

elif test_os == "Darwin":
    print("맥입니다.",test_os)

    # 바탕화면 경로 가져오기
    mac_desktop_path = os.path.expanduser("~/Desktop")

    # test 폴더 경로 생성
    mac_test_folder_path = os.path.join(mac_desktop_path, "test")

    # 로그인, 구글 연동 정보등 초시 세팅 정보 엑셀 파일
    info_file_path = os.path.join(mac_test_folder_path, "info.xlsx")

    # 출고 관련 정보(배송 요청) : 원본 파일    
    deal_out_excel_upload_path = os.path.join(mac_test_folder_path, "auto_out_org_file.xlsx")
    # deal_out_excel_upload_path = r"auto_out_org_file.xlsx" # Excel 파일 경로 설정

    # 출고에 전달할 정보 : 원본 파일
    deal_in_to_out_excel_path = os.path.join(mac_test_folder_path, "info_in_to_out.xlsx")
    # deal_in_to_out_excel_path = r"info_in_to_out.xlsx" # Excel 파일 경로 설정

    #chrone 경로 - info 엑셀 파일에 작성
    # chrome_path = 'Applications/Google Chrome 2'

else:
    print("윈도우나 맥이 아닙니다.",test_os)

# info_file_path = 'C:\\test\\info.xlsx'

try:
    info_workbook = openpyxl.load_workbook(info_file_path)
except FileNotFoundError:
    # 파일이 없을 경우 새로 생성
    info_workbook = openpyxl.Workbook()


if 'Sheet1' in info_workbook.sheetnames:
    info_sheet = info_workbook.active
else:
    info_sheet = info_workbook.create_sheet('Sheet1')


# info_file = askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "C:\\test\\info.xlsx")])

# info_file = askopenfilename(info_file_path)

# 엑셀 파일 열기
# info_workbook = openpyxl.load_workbook(info_file)

# 로그인 정보 입력 받기
#deal_admin_login_id = info_workbook.active['A1']
#deal_admin_login_password = info_workbook.active['B1']
deal_admin_login_id_cell = info_sheet['A1']
deal_admin_login_password_cell = info_sheet['A2']

deal_seller_login_id_cell = info_sheet['A3']
deal_seller_login_password_cell = info_sheet['A4']

wms_login_id_cell = info_sheet['A5']
wms_login_passWord_cell = info_sheet['A6']

font_color = Font(color='FFFFFF')
# print("0번째", deal_admin_login_id_cell)
# print("0번째", deal_admin_login_password_cell)

##### 어드민 계정 정보 #####
if deal_admin_login_id_cell.value is None:
    deal_admin_login_id_input = (input("어드민 로그인 ID: "))
    deal_admin_login_id_cell.value = deal_admin_login_id_input
    info_sheet['A1'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_id = deal_admin_login_id_input
    # print("1번째", deal_admin_login_id)
else:
    deal_admin_login_id = deal_admin_login_id_cell.value
    #print("2번째", deal_admin_login_id)


if deal_admin_login_password_cell.value is None:
    deal_admin_login_password_input = (input("어드민 로그인 비밀번호: "))
    deal_admin_login_password_cell.value = deal_admin_login_password_input
    info_sheet['A2'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_password = deal_admin_login_password_input
    #print("1번째", deal_admin_login_password)

else:
    deal_admin_login_password = deal_admin_login_password_cell.value
    #print("2번째", deal_admin_login_password)

##### 셀러 계정 정보 #####
if deal_seller_login_id_cell.value is None:
    deal_seller_login_id_input = (input("셀러 로그인 ID: "))
    deal_seller_login_id_cell.value = deal_seller_login_id_input
    info_sheet['A3'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_id = deal_seller_login_id_input
    # print("1번째", deal_seller_login_id)
else:
    deal_seller_login_id = deal_seller_login_id_cell.value
    #print("2번째", deal_seller_login_id)


if deal_seller_login_password_cell.value is None:
    deal_seller_login_password_input = (input("셀러 로그인 비밀번호: "))
    deal_seller_login_password_cell.value = deal_seller_login_password_input
    info_sheet['A4'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_password = deal_seller_login_password_input
    #print("1번째", deal_seller_login_password)

else:
    deal_seller_login_password = deal_seller_login_password_cell.value
    #print("2번째", deal_seller_login_password)


##### WMS 계정 정보 #####
if wms_login_id_cell.value is None:
    wms_login_id_input = (input("WMS 로그인 ID: "))
    wms_login_id_cell.value = wms_login_id_input
    info_sheet['A5'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_id = wms_login_id_input
    # print("1번째", wms_login_id)
else:
    wms_login_id = wms_login_id_cell.value
    #print("2번째", wms_login_id)


if wms_login_passWord_cell.value is None:
    wms_login_passWord_input = (input("WMS 로그인 비밀번호: "))
    wms_login_passWord_cell.value = wms_login_passWord_input
    info_sheet['A6'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_passWord = wms_login_passWord_input
    #print("1번째", wms_login_passWord)

else:
    wms_login_passWord = wms_login_passWord_cell.value
    #print("2번째", wms_login_passWord)


# MAC OS 일 경우 크롬 경로 지정
if test_os == "Darwin":
    chrome_path_cell = info_sheet['A12']

    if chrome_path_cell.value is None:
        chrome_path_input = (input("크롬 실행 파일 경로를 입력하세요: "))
        chrome_path_cell.value = chrome_path_input
        info_sheet['A12'].font = font_color
        info_workbook.save(info_file_path)
        chrome_path = chrome_path_input
        #print("1번째", wms_login_passWord)
        
    else:
        chrome_path = chrome_path_cell.value
    #print("2번째", wms_login_passWord)
     


#########################################################################################################################
# 구글 시트 연동

# 구글 json 경로
json_file_name_cell = info_sheet['A7']

if json_file_name_cell.value is None:
    json_file_name = filedialog.askopenfilename()    
    json_file_name_cell.value = json_file_name
    info_sheet['A7'].font = font_color
    info_workbook.save(info_file_path)
    # print("1번째", deal_admin_login_id)
else:
    json_file_name = json_file_name_cell.value
    #print("2번째", deal_admin_login_id)

# json_file_name = 'C:\\auto_json\\fulfillment-371610-dee41b117bdb.json'	# 구글 시트 jSON



scope = [
'https://spreadsheets.google.com/feeds',
'https://www.googleapis.com/auth/drive',
]

gc = gspread.service_account(filename=json_file_name)


# 구글 테스트 시나리오 엑셀 주소
google_url_cell = info_sheet['A8']

if google_url_cell.value is None:
    google_url_input = (input("구글 테스트 시나리오 엑셀 주소: "))
    google_url_cell.value = google_url_input
    info_sheet['A8'].font = font_color
    info_workbook.save(info_file_path)
    google_url = google_url_input
    #print("1번째", deal_seller_login_password)

else:
    google_url = google_url_cell.value
    #print("2번째", deal_seller_login_password)

# google_url = 'https://docs.google.com/spreadsheets/d/1fMf-pNUosGPMJ6evQJLTRRhAGpqZwEIhmC48wxLXHMQ/edit#gid=651282265' # 테스트 시나리오 엑셀 주소


# 구글 API 사용 이메일
google_email_cell = info_sheet['A9']

if google_email_cell.value is None:
    google_email_input = (input("구글 API 사용 이메일: "))
    google_email_cell.value = google_email_input
    info_sheet['A9'].font = font_color
    info_workbook.save(info_file_path)
    google_email = google_email_input
    #print("1번째", deal_seller_login_password)
else:
    google_email = google_email_cell.value

# google_email = 'client_email: fulfillment-test@fulfillment-371610.iam.gserviceaccount.com'

google_doc = gc.open_by_url(google_url)


# 구글 시트
google_sheet_cell = info_sheet['A10']

if google_sheet_cell.value is None:
    google_sheet_input = (input("구글 시트: "))
    google_sheet_cell.value = google_sheet_input
    info_sheet['A10'].font = font_color
    info_workbook.save(info_file_path)
    google_sheet = google_sheet_input
    #print("1번째", deal_seller_login_password)
else:
    google_sheet = google_sheet_cell.value

# google_sheet = '1.정상입고-사입앱' # 구글 시트


google_sheet = google_doc.worksheet(google_sheet)


# 출고 관련 정보(배송 요청) : 원본 파일
#deal_out_excel_upload_path = r"c:\test\auto_out_org_file.xlsx" # Excel 파일 경로 설정

# 출고에 전달할 정보 : 원본 파일
# deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정

#########################################################################################################################
# 함수 선언
def wms_side_columns_select_sendkeys(driver, css_selector):
    wms_side_columns_input_field = driver.find_element(By.CSS_SELECTOR, css_selector)
    time.sleep(2)

    # 플랫폼에 따라 키 조합 결정
    if platform.system() == 'Darwin':
        # macOS
        time.sleep(2)
        wms_side_columns_input_field.send_keys(Keys.COMMAND + 'a')
    else:
        # Windows 또는 Linux
        time.sleep(2)
        wms_side_columns_input_field.send_keys(Keys.CONTROL + 'a')
                
        time.sleep(2)
        wms_side_columns_input_field.send_keys(Keys.DELETE)
        

def deal_search_sendkeys(driver, ID):
    deal_search_field = driver.find_element(By.ID, ID)
    time.sleep(2)

    # 플랫폼에 따라 키 조합 결정
    if platform.system() == 'Darwin':
        # macOS
        time.sleep(2)
        deal_search_field.send_keys(Keys.COMMAND + 'a')
    else:
        # Windows 또는 Linux
        time.sleep(2)
        deal_search_field.send_keys(Keys.CONTROL + 'a')
                
        time.sleep(2)
        deal_search_field.send_keys(Keys.DELETE)


#########################################################################################################################
# 사입 요청 엑셀 파일 -> 사입앱 -> 사입 수량, 도매 단가, 최종 금액 변경에 대한 정보 입력

buyer_workbook = openpyxl.load_workbook(deal_test_saip_excel_upload)
buyer_worsheet = buyer_workbook.active

buyer_purchasedCount_01 = '' # 사입을 위한 모두 사입 수량
buyer_purchasedCount_02 = ''
buyer_purchasedCount_03 = ''

deal_prace_1 = '' # 사입 요청한 첫번째 SKU 도매 단가
deal_prace_2 = '' # 사입 요청한 두번째 SKU 도매 단가
deal_prace_3 = '' # 사입 요청한 세번째 SKU 도매 단가

    
buyer_purchasedCount_01 = buyer_worsheet['I4'].value
buyer_purchasedCount_02 = buyer_worsheet['I5'].value
buyer_purchasedCount_03 = buyer_worsheet['I6'].value

deal_prace_1 = buyer_worsheet['H4'].value
deal_prace_2 = buyer_worsheet['H5'].value
deal_prace_3 = buyer_worsheet['H6'].value

# 상세한 입고 작업(구글 로그 확인)
if test_google_log == 9:
    # # 사입을 위한 모두 사입 수량 01 -> 구글 시트에 업데이트
    google_sheet.update_acell('H12', buyer_purchasedCount_01)
    google_sheet.update_acell('H14', buyer_purchasedCount_02)
    google_sheet.update_acell('H16', buyer_purchasedCount_03)

    google_sheet.update_acell('H11', deal_prace_1)
    google_sheet.update_acell('H13', deal_prace_2)
    google_sheet.update_acell('H15', deal_prace_3)


#########################################################################################################################

#########################################################################################################################
# 크롭 탭 3개 실행
#chrome_options = Options()
#chrome_options.add_argument('--start-maximized')

#driver = webdriver.Chrome(chrome_options=chrome_options)
options = Options()
options.add_argument('--start-maximized')

if test_os == "Windows":
    print("윈도우입니다.", test_os)
    driver = webdriver.Chrome(options=options)
    
elif test_os == "Darwin":
    print("맥입니다.",test_os)
    
    #chrone 경로
    chrome_path = 'Applications/Google Chrome'
    #chrome_path = '/Applications/Google\ Chrome\ 2.app'
    driver = webdriver.Chrome(service=Service(executable_path=chrome_path), options=options)



driver.execute_script('window.open("about:blank", "_blank");')
driver.execute_script('window.open("about:blank", "_blank");')

tabs = driver.window_handles

driver.switch_to.window(tabs[0])
driver.get(deal_seller_url) #신상마켓 소매 -> 딜리버드 진입

driver.maximize_window()

driver.switch_to.window(tabs[1])
driver.get(deal_admin_url)

#driver.set_window_size(6000, 1024)
driver.switch_to.window(tabs[2])
driver.get(wms_url)

time.sleep(2)
# wms 각 항목 -> 조회 -> 리스트의 스크롤이 생기면서 데이터 로드의 어려움(현재 보여지는 화면의 데이터만 가져옴)
# pyautogui을 사용 -> wms 로그인 화면에서 글꼴 축소 -> 항목 조회 -> 리스트의 모든 데이터가 보이게 됨

action = ActionChains(driver)

# hihida 221230
#########################################################################################################################
########### 어드민 -> 딜리버드 셀러 이동 ##########
#########################################################################################################################
driver.switch_to.window(tabs[0])


#어드민 로그인 진행
#driver.find_element(By.ID, 'ssm_admin_email').send_keys(deal_admin_login_id)
#driver.find_element(By.ID, 'ssm_admin_password').send_keys(deal_admin_login_password)
#driver.find_element(By.NAME, 'commit').click()

#어드민에서 테스트 셀러 계정 검색 -> 딜리버드 파트너 센터 이동
#driver.find_element(By.ID, 'inline_search').send_keys(deal_seller_id)
#driver.find_element(By.ID, 'inline_search').send_keys(Keys.ENTER)
#element = driver.find_element(By.ID,'sel_date_1month')# 라디오 버튼 클릭
#driver.execute_script("arguments[0].click();", element)

#driver.implicitly_wait(5)
#셀러 ID 클릭
#time.sleep(5)
#driver.find_element(By.XPATH, '//*[@id="purchasesList"]/tbody/tr[1]/td[3]/a').click()


##신상마켓 소매 -> 딜리버드 이동
#신상마켓 로그인
time.sleep(3)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/header/div/div[2]/div[3]').click() # 로그인 버튼(페이지 상단)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[1]/input').send_keys(deal_seller_login_id) # 모달 / ID 입력
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[2]/input').send_keys(deal_seller_login_password) # 모달 / 비번 입력
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[3]/div[2]/div/button').click() # 모달 / 로그인 버튼
print("##########")
print("신상마켓 소매 로그인")


time.sleep(3)
# 팝업 광고 있을 경우 닫기 클릭
try:
    print("팝업 광고 시작")
    deal_test_result = driver.find_element(By.CLASS_NAME, "close-button")
    #deal_test_result = driver.find_element(By.CSS_SELECTOR, 'div[class="button.close-button.full-btn"]')
    #deal_test_result = driver.find_element(By.CSS_SELECTOR, ".popup-footer__button-area .close-button")
    print("팝업 광고 시작1")
    deal_test_result.click()
    print("팝업 광고 종료1")
    
    time.sleep(2)
    
    deal_test_result = driver.find_element(By.CLASS_NAME, "close-button")
    
    print("팝업 광고 시작2")
    deal_test_result.click()
    print("팝업 광고 종료2")
    
    time.sleep(2)
    
    deal_test_result = driver.find_element(By.CLASS_NAME, "close-button")
    
    print("팝업 광고 시작3")
    deal_test_result.click()
    print("팝업 광고 종료3")
    
    time.sleep(2)
    
    deal_test_result = driver.find_element(By.CLASS_NAME, "close-button")
    
    print("팝업 광고 시작4")
    deal_test_result.click()
    print("팝업 광고 종료4")
    
except:
    pass


#딜리버드 바로가기 클릭
time.sleep(3)

try:
    driver.find_element(By.XPATH,'//*[@id="app"]/div[1]/div[1]/div[1]/div/div[2]/div[1]/div/span').click()
    time.sleep(1)
except:
       
    deal_test_result = driver.find_element(By.CSS_SELECTOR, ".flex-center-center.cursor-pointer.hover\\:text-gray-100")
    deal_test_result.click()
       
    time.sleep(5)
# driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[1]/div/span').click()
print("##########")
print("딜리버드 이동")

