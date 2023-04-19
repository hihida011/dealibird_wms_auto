import selenium.webdriver.support.ui as ui

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog

from openpyxl.styles import Font

import pyautogui
import time
import sys
import json
import re
import openpyxl
import subprocess


import gspread
from oauth2client.service_account import ServiceAccountCredentials

import requests
from bs4 import BeautifulSoup



#########################################################################################################################
# 사입 요청 파일 다운로드
# https://docs.google.com/spreadsheets/d/19X1duCg7N2npHQHGu_pPcDaji_9pDWdI/edit#gid=1830395100
# 해당 엑셀 파일을 c:\test\ 에 저장



"""
##################### 중요!!!!!
# 테스트 전 복사해서 선언해야 함 deal_test_saip_excel_upload, buyer_wsIdx_name

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화11.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23163' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화12.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23164' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화13.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23165' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화14.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23166' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화15.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23167' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화16.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23168' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화17.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23169' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화18.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23170' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화19.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23171' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값

"""
# deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
# buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값


# 불러오기 창 생성
root = tk.Tk()
root.withdraw()
# excel_upload_file_path = filedialog.askopenfilename()
deal_test_saip_excel_upload = filedialog.askopenfilename()

# 선택된 파일 열기

# deal_test_saip_excel_upload = openpyxl.load_workbook(excel_upload_file_path)
# org_workbook = openpyxl.load_workbook(excel_upload_file_path)
# org_worsheet = org_workbook.active

# 사입 요청한 도매 wsIdx 값
buyer_wsIdx_name = input("Enter 도매 wsIndex number: ")


#########################################################################################################################
# 테스트 기본 설정
#########################################################################################################################
# 자동화 실행 시 기본 정보 로드 : ex)계정 정보 입력, 구글 접속 정보
deal_admin_login_id = ''
deal_admin_login_password = ''
deal_admin_url = 'https://dealibird.qa.sinsang.market/ssm_admins/sign_in'
deal_seller_login_id = ''
deal_seller_login_password = ''
deal_seller_url = 'https://vat.qa.sinsang.market/'

# WMS 테스트 기본 설정
wms_login_id = '' 									# WMS 로그인 ID
wms_login_passWord = ''  						# WMS 로그인 비번
wms_url = 'https://matrix-web.qa.sinsang.market/signin'

info_file_path = 'C:\\test\\info.xlsx'
try:
    info_workbook = openpyxl.load_workbook(info_file_path)
except FileNotFoundError:
    # 파일이 없을 경우 새로 생성
    info_workbook = openpyxl.Workbook()


if 'Sheet1' in info_workbook.sheetnames:
    info_sheet = info_workbook.active
else:
    info_sheet = info_workbook.create_sheet('Sheet1')


# info_file = askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "C:\\test\\info.xlsx")])

# info_file = askopenfilename(info_file_path)

# 엑셀 파일 열기
# info_workbook = openpyxl.load_workbook(info_file)

# 로그인 정보 입력 받기
#deal_admin_login_id = info_workbook.active['A1']
#deal_admin_login_password = info_workbook.active['B1']
deal_admin_login_id_cell = info_sheet['A1']
deal_admin_login_password_cell = info_sheet['A2']

deal_seller_login_id_cell = info_sheet['A3']
deal_seller_login_password_cell = info_sheet['A4']

wms_login_id_cell = info_sheet['A5']
wms_login_passWord_cell = info_sheet['A6']

font_color = Font(color='FFFFFF')
# print("0번째", deal_admin_login_id_cell)
# print("0번째", deal_admin_login_password_cell)

##### 어드민 계정 정보 #####
if deal_admin_login_id_cell.value is None:
    deal_admin_login_id_input = (input("어드민 로그인 ID: "))
    deal_admin_login_id_cell.value = deal_admin_login_id_input
    info_sheet['A1'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_id = deal_admin_login_id_input
    # print("1번째", deal_admin_login_id)
else:
    deal_admin_login_id = deal_admin_login_id_cell.value
    #print("2번째", deal_admin_login_id)


if deal_admin_login_password_cell.value is None:
    deal_admin_login_password_input = (input("어드민 로그인 비밀번호: "))
    deal_admin_login_password_cell.value = deal_admin_login_password_input
    info_sheet['A2'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_password = deal_admin_login_password_input
    #print("1번째", deal_admin_login_password)

else:
    deal_admin_login_password = deal_admin_login_password_cell.value
    #print("2번째", deal_admin_login_password)

##### 셀러 계정 정보 #####
if deal_seller_login_id_cell.value is None:
    deal_seller_login_id_input = (input("셀러 로그인 ID: "))
    deal_seller_login_id_cell.value = deal_seller_login_id_input
    info_sheet['A3'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_id = deal_seller_login_id_input
    # print("1번째", deal_seller_login_id)
else:
    deal_seller_login_id = deal_seller_login_id_cell.value
    #print("2번째", deal_seller_login_id)


if deal_seller_login_password_cell.value is None:
    deal_seller_login_password_input = (input("셀러 로그인 비밀번호: "))
    deal_seller_login_password_cell.value = deal_seller_login_password_input
    info_sheet['A4'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_password = deal_seller_login_password_input
    #print("1번째", deal_seller_login_password)

else:
    deal_seller_login_password = deal_seller_login_password_cell.value
    #print("2번째", deal_seller_login_password)


##### WMS 계정 정보 #####
if wms_login_id_cell.value is None:
    wms_login_id_input = (input("WMS 로그인 ID: "))
    wms_login_id_cell.value = wms_login_id_input
    info_sheet['A5'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_id = wms_login_id_input
    # print("1번째", wms_login_id)
else:
    wms_login_id = wms_login_id_cell.value
    #print("2번째", wms_login_id)


if wms_login_passWord_cell.value is None:
    wms_login_passWord_input = (input("WMS 로그인 비밀번호: "))
    wms_login_passWord_cell.value = wms_login_passWord_input
    info_sheet['A6'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_passWord = wms_login_passWord_input
    #print("1번째", wms_login_passWord)

else:
    wms_login_passWord = wms_login_passWord_cell.value
    #print("2번째", wms_login_passWord)




# 출고 관련 정보(배송 요청) : 원본 파일
deal_out_excel_upload_path = r"c:\test\auto_out_org_file.xlsx" # Excel 파일 경로 설정

# 출고에 전달할 정보 : 원본 파일
deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정



#########################################################################################################################
# 사입 요청 엑셀 파일 -> 사입앱 -> 사입 수량, 도매 단가, 최종 금액 변경에 대한 정보 입력

buyer_workbook = openpyxl.load_workbook(deal_test_saip_excel_upload)
buyer_worsheet = buyer_workbook.active

buyer_purchasedCount_01 = '' # 사입을 위한 모두 사입 수량
buyer_purchasedCount_02 = ''
buyer_purchasedCount_03 = ''

deal_prace_1 = '' # 사입 요청한 첫번째 SKU 도매 단가
deal_prace_2 = '' # 사입 요청한 두번째 SKU 도매 단가
deal_prace_3 = '' # 사입 요청한 세번째 SKU 도매 단가

    
buyer_purchasedCount_01 = buyer_worsheet['I4'].value
buyer_purchasedCount_02 = buyer_worsheet['I5'].value
buyer_purchasedCount_03 = buyer_worsheet['I6'].value

deal_prace_1 = buyer_worsheet['H4'].value
deal_prace_2 = buyer_worsheet['H5'].value
deal_prace_3 = buyer_worsheet['H6'].value

# # 사입을 위한 모두 사입 수량 01 -> 구글 시트에 업데이트
#google_sheet.update_acell('H12', buyer_purchasedCount_01)
#google_sheet.update_acell('H14', buyer_purchasedCount_02)
#google_sheet.update_acell('H16', buyer_purchasedCount_03)

#google_sheet.update_acell('H11', deal_prace_1)
#google_sheet.update_acell('H13', deal_prace_2)
#google_sheet.update_acell('H15', deal_prace_3)


#########################################################################################################################

#########################################################################################################################
# 크롭 탭 3개 실행
chrome_options = Options()
chrome_options.add_argument('--start-maximized')

driver = webdriver.Chrome(chrome_options=chrome_options)
driver.execute_script('window.open("about:blank", "_blank");')
driver.execute_script('window.open("about:blank", "_blank");')

tabs = driver.window_handles

driver.switch_to.window(tabs[0])
driver.get(deal_seller_url) #신상마켓 소매 -> 딜리버드 진입

driver.maximize_window()

driver.switch_to.window(tabs[1])
driver.get(deal_admin_url)

#driver.set_window_size(6000, 1024)
driver.switch_to.window(tabs[2])
driver.get(wms_url)

time.sleep(2)
# wms 각 항목 -> 조회 -> 리스트의 스크롤이 생기면서 데이터 로드의 어려움(현재 보여지는 화면의 데이터만 가져옴)
# pyautogui을 사용 -> wms 로그인 화면에서 글꼴 축소 -> 항목 조회 -> 리스트의 모든 데이터가 보이게 됨

action = ActionChains(driver)

# hihida 221230
#########################################################################################################################
########### 어드민 -> 딜리버드 셀러 이동 ##########
#########################################################################################################################
driver.switch_to.window(tabs[0])


#어드민 로그인 진행
#driver.find_element(By.ID, 'ssm_admin_email').send_keys(deal_admin_login_id)
#driver.find_element(By.ID, 'ssm_admin_password').send_keys(deal_admin_login_password)
#driver.find_element(By.NAME, 'commit').click()

#어드민에서 테스트 셀러 계정 검색 -> 딜리버드 파트너 센터 이동
#driver.find_element(By.ID, 'inline_search').send_keys(deal_seller_id)
#driver.find_element(By.ID, 'inline_search').send_keys(Keys.ENTER)
#element = driver.find_element(By.ID,'sel_date_1month')# 라디오 버튼 클릭
#driver.execute_script("arguments[0].click();", element)

#driver.implicitly_wait(5)
#셀러 ID 클릭
#time.sleep(5)
#driver.find_element(By.XPATH, '//*[@id="purchasesList"]/tbody/tr[1]/td[3]/a').click()


##신상마켓 소매 -> 딜리버드 이동
#신상마켓 로그인
time.sleep(3)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/header/div/div[2]/div[3]').click() # 로그인 버튼(페이지 상단)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[1]/input').send_keys(deal_seller_login_id) # 모달 / ID 입력
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[2]/input').send_keys(deal_seller_login_password) # 모달 / 비번 입력
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[3]/div[2]/div/button').click() # 모달 / 로그인 버튼
print("##########")
print("신상마켓 소매 로그인")



#딜리버드 바로가기 클릭
time.sleep(3)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[1]/div/span').click()
print("##########")
print("딜리버드 이동")


#########################################################################################################################
#### 딜리버드 -> 사입 요청 #####
time.sleep(5)

try:
    driver.find_element(By.XPATH,'//*[@id="navbarSupportedContent"]/ul/li[2]/a').click()
    time.sleep(5)
except:
    deal_link = driver.find_element(By.LINK_TEXT, ("사입요청"))
    deal_link.click()
    time.sleep(5)
    
print("##########")
print("사입 요청 시작")


# 기존에 등록했던 사입 요청 리스트가 있다면 입력 초기화 하고 새로 입력
deal_test_result = driver.find_element(By.XPATH,'//*[@id="purchase_totalCount"]') # 페이지 중간 왼쪽 -> 사입 요청 : X 값
deal_test_result_check = deal_test_result.text # 사입 요청 값에서 테스트 값을 저장

if deal_test_result_check != "-": # 사입 요청 건수가 0건일 경우 실행 되지 않음
    deal_table_count = int(deal_test_result_check) # 텍스트 값을 int형으로 형변환
    
    if deal_table_count > 0: # 사입 요청 건 수가 1건 이상 있을 경우
        driver.find_element(By.XPATH, '//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[1]').click() # 페이지 중간의 [입력초기화] 버튼
        time.sleep(3)
    
        driver.find_element(By.XPATH, '/html/body/div[5]/div/div[3]/button[3]').click() # 얼럿 -> 모두 초기화 ~ [예] 버튼
        time.sleep(2)

driver.find_element(By.XPATH, '//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[6]').click() # 페이지 중간의 [엑셀 업로드] 버튼
time.sleep(2)

#파일 업로드
up_load_file = driver.find_element(By.XPATH, '//*[@id="excel_file"]') # 모달 / 엑셀 업로드 양식 선태 -> 엑셀 파일 선택 Browse 버튼
up_load_file.send_keys(deal_test_saip_excel_upload)
driver.find_element(By.NAME, 'commit').click() # [저장] 버튼

print("##########")
print("파일 업로드 완료")
time.sleep(5)

# 사입 요청 리스트(테이블) 체크
deal_table = driver.find_element(By.XPATH, '//*[@id="purchasesList"]') # 리스트(테이블) 전체 경로
deal_tbody = deal_table.find_element(By.TAG_NAME,'tbody')
deal_list_count = 0
deal_saib_do_store_name = "" # 도매 매장명


for tr in deal_tbody.find_elements(By.TAG_NAME,'tr'):
    for td in tr.find_elements(By.TAG_NAME,'td'):
        if deal_list_count == 13: # 테스트 데이터를 위해 엑셀에 업로드, 도매 매장명 # # 도매 명  -> 구글 시트에 업데이트
            deal_saib_do_store_name = td.get_attribute("innerText")
            break
        
        deal_list_count += 1




driver.find_element(By.XPATH, '//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[9]').click() # 페이지 중간 오른쪽 [사입 요청하기] 버튼
time.sleep(2)


# 사입 요청 버튼
driver.find_element(By.XPATH, '/html/body/div[5]/div/div[3]/button[3]').click() # 얼럿 / X건의 상품을~~ 진행하시겠습니까? -> [네] 버튼
time.sleep(5)
print("##########")
print("사입 요청 완료")


# 결제 정보 모달
driver.find_element(By.ID, 'method_SINSANGPOINT').click() # 모달 / 결제 수단 선택 -> 신상캐시 버튼
time.sleep(2)
#google_sheet.update_acell('I17', 'OK')
print(" 17 PASS")
print(" PASS")

try:
    driver.find_element(By.XPATH,'//*[@id="confirmCollapse"]/div[2]/div/label').click() # 모달 -> 이용 약관 -> 전체 동의합니다. 버튼
except:
    print("이용 약관 xpath 값 오류 except 진입")
    deal_test_result = driver.find_element(By.id, "policyAllCheck")
    deal_test_result.click()


try:
    driver.find_element(By.XPATH, '//*[@id="payment_button"]').click() # 모달 -> [결제하기] 버튼
except:
    print("모달 -> [결제하기] 버튼 xpath 값 오류 except 진입")
    deal_test_result = driver.find_element(By.id, "payment_button")
    deal_test_result.click()    
    

time.sleep(10)

print("##########")
print("결제 완료")
#google_sheet.update_acell('I18', 'OK')
print("18 PASS")

#########################################################################################################################
# 결제 완료 후 사입 요청 페이지
# 중요 hihida 딜리버드 주문번호 저장 / 사입 요청 번호
deal_wms_purchase_number = driver.find_element(By.XPATH, '//*[@id="page-wrapper"]/div[2]/div[2]/div/div/div[1]/div[1]/h4/span')
deal_wms_purchase_number = deal_wms_purchase_number.text

# 딜리버드 주문번호 -> 구글 시트에 업데이트
#google_sheet.update_acell('H43', deal_wms_purchase_number)
##google_sheet.update_acell('H50', deal_wms_purchase_number)
##google_sheet.update_acell('H54', deal_wms_purchase_number)
##google_sheet.update_acell('H78', deal_wms_purchase_number)
##google_sheet.update_acell('H102', deal_wms_purchase_number)
##google_sheet.update_acell('H118', deal_wms_purchase_number)


# 중요 hihida 딜리버드 상품코드 저장
deal_product_id_1 = driver.find_element(By.XPATH, '//*[@id="purchasesDetail"]/tbody/tr[1]/td[2]')
deal_product_id_2 = driver.find_element(By.XPATH, '//*[@id="purchasesDetail"]/tbody/tr[2]/td[2]')
deal_product_id_3 = driver.find_element(By.XPATH, '//*[@id="purchasesDetail"]/tbody/tr[3]/td[2]')

deal_product_id_1 = deal_product_id_1.text
deal_product_id_2 = deal_product_id_2.text
deal_product_id_3 = deal_product_id_3.text

# 딜리버드 상품코드  -> 구글 시트에 업데이트
##google_sheet.update_acell('H30', deal_product_id_1)
#google_sheet.update_acell('H103', deal_product_id_1)
##google_sheet.update_acell('H119', deal_product_id_1)
##google_sheet.update_acell('H126', deal_product_id_1)

##google_sheet.update_acell('H34', deal_product_id_2)
#google_sheet.update_acell('H106', deal_product_id_2)
##google_sheet.update_acell('H121', deal_product_id_2)
##google_sheet.update_acell('H130', deal_product_id_2)

##google_sheet.update_acell('H38', deal_product_id_3)
#google_sheet.update_acell('H109', deal_product_id_3)
##google_sheet.update_acell('H123', deal_product_id_3)
##google_sheet.update_acell('H134', deal_product_id_3)




#########################################################################################################################
########## 어드민 사입마감 처리 ##########
#########################################################################################################################

driver.switch_to.window(tabs[1])


#어드민 로그인 진행
time.sleep(10)
driver.find_element(By.ID, 'ssm_admin_email').send_keys(deal_admin_login_id)
driver.find_element(By.ID, 'ssm_admin_password').send_keys(deal_admin_login_password)
driver.find_element(By.NAME, 'commit').click()
time.sleep(2)
print("##########")
print("어드민 로그인 완료")

deal_admin_saip_end_url = 'https://dealibird.qa.sinsang.market/admin/pps/purchase_schedules/manual_transfer/' + deal_wms_purchase_number
# admin_response = requests.post(url=deal_admin_saip_end_url) # 사입 마감 처리
driver.get(deal_admin_saip_end_url) # 사입 마감 처리



print("##########")
# print("사입 마감 처리 완료", admin_response)
print("사입 마감 처리 완료")
# hihida 221230
#google_sheet.update_acell('I19', 'OK')
print("19 PASS")
time.sleep(10)


#########################################################################################################################
########### 사입앱 로그인 ##########
#########################################################################################################################
buyer_login_url = 'https://buyer.qa.sinsang.market/api/v1/session' # 사입 로그인 URL
buyer_login_header = {'Content-Type' : 'application/json', "User-Agent" : "Mozilla/5.0"} # 로그인 시 헤더
buyer_login_data = {
    'password':'1234',
    'user':'qa_smkim'
} # 로그인 시 Body 정보 : 로그인 계정

buyer_response = requests.post(url=buyer_login_url, headers=buyer_login_header, params=buyer_login_data) # 로그인 시도
print("사입앱 로그인 성공\n", buyer_response)

buyer_login_content = buyer_response.content # 로그인 후 리턴되는 값(여러 정보가 있음)
buyer_login_content_data = json.loads(buyer_login_content) # JSON 문자열을 Python 객체로 변환
buyer_login_content_accesstoken = buyer_login_content_data["content"]["accessToken"] # accessToken 저장




#########################################################################################################################
########### 사입 리스트 상세 : 사입 예정 SKU의 ID 가져오기 ##########

buyer_login_accesstoken_header = {'Content-Type': 'application/json',
#'access-token': login_accesstoken,
'access-token': buyer_login_content_accesstoken,
'User-Agent': 'Mozilla/5.0',
'Cache-Control': 'no-cache',
'Accept': '*/*',
'Host': 'buyer.qa.sinsang.market',
'Accept-Encoding': 'gzip, deflate, br',
'Connection': 'keep-alive'} # accessToken을 가지고 사입 리스트 상세 조회


# id_search_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_detail?wsIdx=23124&orderType=purchase&warehouse=B1'

# buyer_id_search_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_detail?wsIdx='+ buyer_wsIdx_name + '&orderType=purchase&warehouse=B1' # 사입 리스트 상세 조회 URL
buyer_id_search_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_detail?wsIdx='+ buyer_wsIdx_name + '&orderType=purchase' # 사입 리스트 상세 조회 URL
print(buyer_wsIdx_name)

buyer_response = requests.get(url=buyer_id_search_url, headers=buyer_login_accesstoken_header) # 사입 리스트 상세 조회 시도
print("사입 리스트 상세 조회 성공\n", buyer_response)

buyer_id_search_content = buyer_response.content# 조회 후 리턴되는 값(여러 정보가 있음)

buyer_id_search_content_data = json.loads(buyer_id_search_content) # JSON 문자열을 Python 객체로 변환

buyer_id_search_content_ID_data = [] # 여러개의 ID 정보 저장을 위한 배열
buyer_id_search_int = int(0) # 배열 Len 체크
for product in buyer_id_search_content_data["content"]["products"]:
    buyer_id_search_content_ID_data.append(product["id"]) # id 정보를 배열(id_search_content_ID_data)에 저장
    print(buyer_id_search_int, "번째 ID는", buyer_id_search_content_ID_data[buyer_id_search_int])
    buyer_id_search_int = buyer_id_search_int +1

buyer_id_search_int = buyer_id_search_int -1 # 최종 배열 길이 체크


#########################################################################################################################
########### 사입 상품 옵션 저장 : SKU의 ID로 모두 사입으로 전송 ##########

buyer_order_status_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_status' # 사입 상태 전송 URL
                    
buyer_order_status_data = {
    "orderType" : "purchase", # 신규 주문
	"items": [
		{
            "id": buyer_id_search_content_ID_data[0],
            # "purchasedCount" : 5
            "purchasedCount" : buyer_purchasedCount_01
		},
        {
            "id": buyer_id_search_content_ID_data[1],
            "purchasedCount" : buyer_purchasedCount_02
            # "purchasedCount" : 8
		},
        {
            "id": buyer_id_search_content_ID_data[2],
            "purchasedCount" : buyer_purchasedCount_03
            # "purchasedCount" : 10
		}
	]
} # 로그인 시 Body 정보 : 로그인 계정


buyer_response = requests.post(url=buyer_order_status_url, headers=buyer_login_accesstoken_header, data=json.dumps(buyer_order_status_data)) # 사입 상태 전송 시도

print("사입앱 사입 성공 전달 성공\n", buyer_response)

#google_sheet.update_acell('I16', 'OK')
print("16 PASS")


#########################################################################################################################
##어드민 -> WMS 이동
#########################################################################################################################
driver.switch_to.window(tabs[2])


# 테스트를 위한 임시 저장 hihida
#deal_wms_purchase_number = "19674"

# wms 각 항목 -> 조회 -> 리스트의 스크롤이 생기면서 데이터 로드의 어려움(현재 보여지는 화면의 데이터만 가져옴)
# pyautogui을 사용 -> wms 로그인 화면에서 글꼴 축소 -> 항목 조회 -> 리스트의 모든 데이터가 보이게 됨
#pyautogui_count = 0
#while pyautogui_count < 8:
#    pyautogui.hotkey('ctrl', '-') # 화면 축소
#    pyautogui_count = pyautogui_count + 1

# time.sleep(130) # 사입 마감 처리 시간 후 로그인 시도 hihida
time.sleep(10) # 테스트 임시


#WMS 로그인 진행
driver.find_element(By.ID, 'login').send_keys(wms_login_id)
driver.find_element(By.ID, 'password').send_keys(wms_login_passWord)
driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/form/div[2]/button').click()
time.sleep(2)
print("##########")
print("wms 로그인 완료")

#google_sheet.update_acell('I25', 'OK')
print("I25 PASS")

## hihida 230102"""
#########################################################################################################################
##### 입고 관리 - 입고 대기 #####
time.sleep(2)
# 입고 관리 -> 입고 대기 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 대기":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break   
    
 

print("##########")
print("입고 관리 - 입고 대기 이동")
#google_sheet.update_acell('I42', 'OK')
print("I42 PASS")
time.sleep(2)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
time.sleep(2)
#google_sheet.update_acell('I43', 'OK') 
print("I43 PASS")

wms_search_list_row_result = driver.find_element(By.CSS_SELECTOR,'div[class="ag-center-cols-container"]')
wms_search_list_row_id_result = wms_search_list_row_result.find_element(By.CSS_SELECTOR, 'div[row-index="0"]') # 리스트 row 마다 ID가져오기
#wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="4"]') # 소봉바코드 : 04
#wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="purchase_barcode"]') # 소봉바코드
wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="purchaseBarcode"]') # 소봉바코드
wms_small_bag_barcode = wms_search_list_row_result_check_1.text # 소봉 바코드 저장
#google_sheet.update_acell('H49', wms_small_bag_barcode)
print("입고 대기 - 소봉바코드   ", wms_small_bag_barcode)

#google_sheet.update_acell('I47', 'OK') 
print("47 PASS")



#########################################################################################################################
##### 입고 관리 - 입고 검수진행 #####
time.sleep(2)
# 입고 관리 -> 입고 검수진행 이동 old
# driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[1]/div/a[3]').click() # 입고 관리 -> 입고 검수진행 이동

# 입고 관리 -> 입고 검수진행 이동(230116)
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 검수진행":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break

time.sleep(2)
print("##########")
print("입고 관리 - 입고 검수진행 이동")
#google_sheet.update_acell('I48', 'OK') 
print("48 PASS")



time.sleep(5)
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='바코드(도매처 소봉, 상품)를 스캔해 주세요']").send_keys(wms_small_bag_barcode) # "바코드 입력"에 입고대기에서 복사한 바코드 입력
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='바코드(도매처 소봉, 상품)를 스캔해 주세요']").send_keys(Keys.ENTER) # 검색 적용
#google_sheet.update_acell('I49', 'OK') 
print("49 PASS")

time.sleep(3)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
#google_sheet.update_acell('I50', 'OK') 
print("50 PASS")




# 전체 바코드 출력하기
time.sleep(5)
print("##########")
print("입고 검수진행 - 전체 바코드 출력) ] 버튼 시작") # [전체 바코드 출력]버튼 클릭한다.

try:
    driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/button').click() # 중간 -> [전체 바코드 출력] 버튼 클릭
except:
    print("try try try try try try")
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeLarge.MuiButton-containedSizeLarge.MuiButtonBase-root')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "전체 바코드 출력":
            #print("클릭 시도\n")
            wms_str_loop.click()
            #print("클릭 완료\n")
            break

#google_sheet.update_acell('I70', 'OK') 
print("70 PASS")
time.sleep(10)


#driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/div[2]/div/div/button').click() # 모달 -> 전체 상품 바코드 출력 -> 진행 클릭
driver.find_element(By.XPATH,'/html/body/div[3]/div[3]/div[2]/button[2]').click() # 모달 -> 전체 상품 바코드 출력 -> 진행 클릭
#google_sheet.update_acell('I71', 'OK') 
print("71 PASS")
time.sleep(10)

# 젙체 바코드 출력 후 시리얼 저장
alert = driver.switch_to.alert 
alert_barcode_all_print_text = alert.text


alert.accept() # 얼럿 확인
print("##########")
print("입고 검수진행 - 전체 바코드 출력) ] 버튼 종료")
#google_sheet.update_acell('I72', 'OK') 
print("72 PASS")
time.sleep(5)



#########################################################################################################################
##### 입고 관리 - 입고 진행현황 #####

# 입고 관리 -> 입고 진행현황 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 진행현황":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break


#google_sheet.update_acell('I73', 'OK') 
print("73 PASS")
time.sleep(2)
print("##########")
print("입고 관리 - 입고 진행현황 이동")

################################################
## 열(컬럼) 메뉴 - 특정 컬럼만 선택
# wms_test_result = driver.find_elements(By.CLASS_NAME, 'ag-side-button-button') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result[0].click() # 중간 오른쪽의 열 컬럼 버튼 클릭

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라짐
# wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-icon.ag-icon-columns') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result.click() # 중간 오른쪽의 열 컬럼 버튼 클릭
# time.sleep(3) # 0[열 컬럼] / 1[필터]

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라져 다시 수정
# <span class =ag-side-button-labe>에서 열(컬럼), 필터 중 열(컬럼) 찾아 버튼 클릭
time.sleep(2)
wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

time.sleep(3)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click() #전체 선택 버튼 클릭해 일단 전체 컬럼 선택
time.sleep(5)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click()#전체 선택 버튼 다시 클릭해 일단 전체 컬럼 선택 해제
time.sleep(5)


# 체크박스 컬럼 선택
wms_test_result = driver.find_element(By.CSS_SELECTOR, "div>[aria-posinset='2']") # 체크박스 컬럼의 유일값을 찾음, aria-posinset='2'
time.sleep(5)


wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
time.sleep(5)

wms_test_result.find_element(By.ID, wms_test_result_chk).click()

#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 체크박스 유일값으로 동적으로 변화되는 aria-describedby ID값 취득
#time.sleep(5)


#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() 
time.sleep(5)



# 주문번호 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("주문번호") # 컬럼 검색 필드 - 주문번호
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합


wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break


################################################



#cell_data = #google_sheet.acell('H74').value 

driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='도매 매장명']").send_keys(deal_saib_do_store_name) # "도매 매장명"에 도매 명 입력
#google_sheet.update_acell('I74', 'OK') 
print("74 PASS")
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='도매 매장명']").send_keys(Keys.ENTER) # 검색 적용
#google_sheet.update_acell('I75', 'OK') 
print("75 PASS")


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
#google_sheet.update_acell('I78', 'OK') 
print("78 PASS")
time.sleep(2)



   
#리스트의 전체 체크 박스 선택
wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-header-cell.ag-header-cell-sortable.ag-focus-managed') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
wms_test_result.find_element(By.ID, wms_test_result_chk).click()
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합
#google_sheet.update_acell('I84', 'OK') 
print("84 PASS")

#google_sheet.update_acell('I90', 'OK')
#google_sheet.update_acell('I96', 'OK')
# hihida 230102"""


print("##########")
print("(입고 진행현황 - 입고확정가능처리 대상상품만 보기) ] 토클 버튼 실행")
try:
    driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/div[2]/div/div/div[4]/div[2]/label[1]/span[2]').click() # 강제입고확정가능만 보기) ] 토클 버튼 실행
except:
    print("try try try try try try")
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiTypography-root.MuiTypography-body1.MuiFormControlLabel-label')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "입고확정가능처리 대상상품만 보기":
            print("클릭 시도 - 입고확정가능처리 대상상품만 보기)n")
            wms_str_loop.click() # 강제 입고 확장 가능 토글
            print("클릭 완료 - 입고확정가능처리 대상상품만 보기\n")
            break    


time.sleep(5)


print("##########")
print("(입고 진행현황 - 선택상품 입고확정가능 처리) ] 버튼 실행")
try:
    driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/div[2]/div/div/div[4]/div[2]/button[1]').click() # (선택 상품 강제 입고 확정) ] 버튼 종료
    #google_sheet.update_acell('I97', 'OK') 
    print("97 PASS")
    time.sleep(3)
    alert = driver.switch_to.alert 
    alert.accept() # 얼럿 확인
    #google_sheet.update_acell('I98', 'OK') 
    print("98 PASS")
    time.sleep(3)
    # alert.dismiss()# 취소
except:
    print("try try try try try try")
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeSmall.MuiButton-containedSizeSmall.MuiButtonBase-root')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "선택상품 입고확정가능 처리":
            print("클릭 시도 - 선택상품 입고확정가능 처리\n")
            wms_str_loop.click()
            #google_sheet.update_acell('I97', 'OK') 
            print("97 PASS")
            time.sleep(3)
            alert = driver.switch_to.alert 
            alert.accept() # 얼럿 확인
            #google_sheet.update_acell('I98', 'OK') 
            print("98 PASS")
            time.sleep(3)
            # alert.dismiss()# 취소
            print("클릭 완료 - 선택상품 입고확정가능 처리\n")
            break
    



try:
    alert = driver.switch_to.alert
    alert.accept() # 얼럿 확인
    time.sleep(3)
except:
    print("try try try try try try")
    pass



########################################################################################################################
##### 입고 관리 - 입고 확정 가능 #####
# 입고 관리 -> 입고 확정 가능 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 확정 가능":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break
        
#google_sheet.update_acell('I99', 'OK') 
print("99 PASS")
time.sleep(2)
print("##########")
print("입고 관리 - 입고 확정 가능 이동")


#cell_data = #google_sheet.acell('H100').value 

driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='검색어를 입력해주세요.']").send_keys(deal_saib_do_store_name) # "도매 매장명"에 도매 명 입력
#google_sheet.update_acell('I100', 'OK') 
print("100 PASS")
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='검색어를 입력해주세요.']").send_keys(Keys.ENTER) # 검색 적용
#google_sheet.update_acell('I101', 'OK') 
print("101 PASS")
time.sleep(2)

# hihida 230102 WMS 오류로 임시 주석
# driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호

time.sleep(2)

   
#리스트의 전체 체크 박스 선택
wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-header-cell.ag-header-cell-sortable.ag-focus-managed') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
time.sleep(5)

wms_test_result.find_element(By.ID, wms_test_result_chk).click()

#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합
#google_sheet.update_acell('I105', 'OK') 
print("105 PASS")
#google_sheet.update_acell('I108', 'OK')
#google_sheet.update_acell('I111', 'OK')

# [입고 확정하기] 버튼 클릭
try:
    driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[5]/div[2]/button').click()
except:
    print("try try try try try try")
    # 232203 class 값 변경
    # wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeMedium.MuiButton-containedSizeMedium.MuiButtonBase-root')
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeSmall.MuiButton-containedSizeSmall.MuiButtonBase-root')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "입고 확정하기":
            print("클릭 시도\n")
            wms_str_loop.click()
            print("클릭 완료\n")
            break

#google_sheet.update_acell('I112', 'OK') 
print("112 PASS")
time.sleep(3)


alert = driver.switch_to.alert 
alert.accept() # 얼럿 확인
#google_sheet.update_acell('I113', 'OK') 
print("113 PASS")
time.sleep(3)
# alert.dismiss()# 취소

try:
    alert = driver.switch_to.alert
    alert.accept() # 얼럿 확인
    time.sleep(3)
except:
    print("try try try try try try")
    pass

print("##########")
print("입고 확정 가능 - (선택 상품 강제 입고 확정) ] 버튼 종료")





#########################################################################################################################
## WMS 이동 -> 딜리버드 이동
#########################################################################################################################
driver.switch_to.window(tabs[0])


#########################################################################################################################
#### 딜리버드 -> 상품 및 재고 #####

try:
    deal_link = driver.find_element(By.LINK_TEXT, ("상품 및 재고"))
    deal_link.click()
    time.sleep(2)
    # driver.find_element(By.XPATH,'//*[@id="navbarSupportedContent"]/ul/li[8]/a').click()
    #google_sheet.update_acell('I125', 'OK') 
    print("125 PASS")
    time.sleep(5)
    print("##########")
    print("딜리버드 상품 및 재고 시작")
except:
    print("try try try try try try")
    ##신상마켓 소매 -> 딜리버드 이동
    #신상마켓 로그인
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/header/div/div[2]/div[3]').click() # 로그인 버튼(페이지 상단)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[1]/input').send_keys(deal_seller_login_id) # 모달 / ID 입력
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[2]/input').send_keys(deal_seller_login_password) # 모달 / 비번 입력
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[3]/div[2]/div/button').click() # 모달 / 로그인 버튼
    print("##########")
    print("신상마켓 소매 로그인")

    #딜리버드 바로가기 클릭
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[1]/div/span').click()
    print("##########")
    print("딜리버드 이동")
    time.sleep(5)
    
    # driver.find_element(By.XPATH,'//*[@id="navbarSupportedContent"]/ul/li[8]/a').click()
    deal_link = driver.find_element(By.LINK_TEXT, ("상품 및 재고"))
    deal_link.click()
    time.sleep(2)    
    #google_sheet.update_acell('I125', 'OK') 
    print("125 PASS")
    time.sleep(5)
    print("##########")
    print("딜리버드상품 및 재고 시작")

# 0126
# 출고 요청 엑셀 파일
deal_wb = openpyxl.load_workbook(deal_out_excel_upload_path)
deal_ws = deal_wb.active

# 출고 요청에게 전달을 위한 변수
deal_sell_product_name_1 = '' # 판매상품명(판매 상품명_자동화상품001)
deal_sell_product_name_2 = '' # 판매상품명(판매 상품명_자동화상품002)
deal_sell_product_name_3 = '' # 판매상품명(판매 상품명_자동화상품003)


# 도매 상품명_자동화001 상품 및 재고 확인
#cell_data = #google_sheet.acell('H126').value # 구글 시트 -> 딜리버드 상품코드
driver.find_element(By.ID,'search_text').send_keys(deal_product_id_1) # 검색어 창에 "딜리버드 상품 코드" 입력
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.ENTER)
#google_sheet.update_acell('I126', 'OK') 
print("126 PASS")
time.sleep(2)

# 0126
# 출고 요청 엑셀 파일에 딜리버드 상품 토드 입력
deal_ws['R3'] = deal_product_id_1
deal_ws['R4'] = deal_product_id_1



# 도매 상품명_자동화002 상품 및 재고 확인
#cell_data = #google_sheet.acell('H130').value # 구글 시트 -> 딜리버드 상품코드
driver.find_element(By.ID,'search_text').send_keys(deal_product_id_2) # 검색어 창에 "딜리버드 상품 코드" 입력
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.ENTER)
#google_sheet.update_acell('I130', 'OK') 
print("130 PASS")
time.sleep(2)

# 0126
# 출고 요청 엑셀 파일에 딜리버드 상품 토드 입력
deal_ws['R5'] = deal_product_id_2
deal_ws['R6'] = deal_product_id_2



# 도매 상품명_자동화003 상품 및 재고 확인
#cell_data = #google_sheet.acell('H134').value # 구글 시트 -> 딜리버드 상품코드
driver.find_element(By.ID,'search_text').send_keys(deal_product_id_3) # 검색어 창에 "딜리버드 상품 코드" 입력
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.ENTER)
#google_sheet.update_acell('I134', 'OK') 
print("134 PASS")
time.sleep(2)

# 0126
# 출고 요청 엑셀 파일에 딜리버드 상품 토드 입력
deal_ws['R7'] = deal_product_id_3
deal_ws['R8'] = deal_product_id_3



# 0126
# 출고 요청 엑셀 파일 저장
deal_out_excel_upload_name_now = datetime.now()


deal_out_excel_upload_name_new = f"자동화_출고요청_QA사입앱자동화_{deal_out_excel_upload_name_now.strftime('%Y%m%d_%H%M%S')}.xlsx"
deal_out_excel_upload_file_path = r"c:\test\\"+deal_out_excel_upload_name_new
deal_wb.save(deal_out_excel_upload_file_path)


# 정상 출고로 보내야 하는 인자값 정리
deal_test_result = driver.find_element(By.XPATH,'//*[@id="page-wrapper"]/div[1]/nav/div/ul/li[1]/strong') # 셀러 이름(대표자)
deal_seller_name = deal_test_result.text




# 출고에 전달할 정보 : 원본 파일
deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정

try:
    info_in_to_out_workbook = openpyxl.load_workbook(deal_in_to_out_excel_path)
except FileNotFoundError:
    # 파일이 없을 경우 새로 생성
    info_in_to_out_workbook = openpyxl.Workbook()


if 'Sheet1' in info_in_to_out_workbook.sheetnames:
    info_in_to_out_sheet = info_in_to_out_workbook.active
else:
    info_in_to_out_sheet = info_in_to_out_workbook.create_sheet('Sheet1')

info_in_to_out_sheet['B1'].value = deal_product_id_1
info_in_to_out_sheet['B2'].value = deal_product_id_2
info_in_to_out_sheet['B3'].value = deal_product_id_3
info_in_to_out_sheet['B4'].value = deal_seller_name
info_in_to_out_sheet['B5'].value = deal_sell_product_name_1
info_in_to_out_sheet['B6'].value = deal_sell_product_name_2
info_in_to_out_sheet['B7'].value = deal_sell_product_name_3

info_in_to_out_workbook.save(deal_in_to_out_excel_path)

print("파일 저장 완료")
print("1. 정상입고 종료")
driver.quit()



# 자동화 정상 입고-사입앱 파일 종료 -> 정상 출고 실행
# 원본
# subprocess.call(['python', '02.AutoTest_OUT_normal.py', json.dumps(file_info)])

# 테스트
# subprocess.call(['python', 'test1.py', json.dumps(file_info)])


#########################
#########################
#########################
#########################
#########################
#########################
#########################
#########################
#########################




#while(True):
#    	pass
 