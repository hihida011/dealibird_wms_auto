import selenium.webdriver.support.ui as ui

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog

from openpyxl.styles import Font

import pyautogui
import time
import sys
import json
import re
import openpyxl
import subprocess


import gspread
from oauth2client.service_account import ServiceAccountCredentials

import requests
from bs4 import BeautifulSoup



#########################################################################################################################
# 사입 요청 파일 다운로드
# https://docs.google.com/spreadsheets/d/19X1duCg7N2npHQHGu_pPcDaji_9pDWdI/edit#gid=1830395100
# 해당 엑셀 파일을 c:\test\ 에 저장



"""
##################### 중요!!!!!
# 테스트 전 복사해서 선언해야 함 deal_test_saip_excel_upload, buyer_wsIdx_name

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화11.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23163' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화12.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23164' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화13.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23165' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화14.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23166' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화15.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23167' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화16.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23168' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화17.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23169' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화18.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23170' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화19.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23171' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값

"""
# deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
# buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값


#########################################################################################################################
# 테스트 기본 설정
#########################################################################################################################
# 자동화 실행 시 기본 정보 로드 : ex)계정 정보 입력, 구글 접속 정보
deal_admin_login_id = ''
deal_admin_login_password = ''
deal_admin_url = 'https://dealibird.qa.sinsang.market/ssm_admins/sign_in'
deal_seller_login_id = ''
deal_seller_login_password = ''
deal_seller_url = 'https://vat.qa.sinsang.market/'

# WMS 테스트 기본 설정
wms_login_id = '' 									# WMS 로그인 ID
wms_login_passWord = ''  						# WMS 로그인 비번
wms_url = 'https://matrix-web.qa.sinsang.market/signin'

info_file_path = 'C:\\test\\info.xlsx'
try:
    info_workbook = openpyxl.load_workbook(info_file_path)
except FileNotFoundError:
    # 파일이 없을 경우 새로 생성
    info_workbook = openpyxl.Workbook()


if 'Sheet1' in info_workbook.sheetnames:
    info_sheet = info_workbook.active
else:
    info_sheet = info_workbook.create_sheet('Sheet1')


# info_file = askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "C:\\test\\info.xlsx")])

# info_file = askopenfilename(info_file_path)

# 엑셀 파일 열기
# info_workbook = openpyxl.load_workbook(info_file)

# 로그인 정보 입력 받기
#deal_admin_login_id = info_workbook.active['A1']
#deal_admin_login_password = info_workbook.active['B1']
deal_admin_login_id_cell = info_sheet['A1']
deal_admin_login_password_cell = info_sheet['A2']

deal_seller_login_id_cell = info_sheet['A3']
deal_seller_login_password_cell = info_sheet['A4']

wms_login_id_cell = info_sheet['A5']
wms_login_passWord_cell = info_sheet['A6']

font_color = Font(color='FFFFFF')
# print("0번째", deal_admin_login_id_cell)
# print("0번째", deal_admin_login_password_cell)

##### 어드민 계정 정보 #####
if deal_admin_login_id_cell.value is None:
    deal_admin_login_id_input = (input("어드민 로그인 ID: "))
    deal_admin_login_id_cell.value = deal_admin_login_id_input
    info_sheet['A1'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_id = deal_admin_login_id_input
    # print("1번째", deal_admin_login_id)
else:
    deal_admin_login_id = deal_admin_login_id_cell.value
    #print("2번째", deal_admin_login_id)


if deal_admin_login_password_cell.value is None:
    deal_admin_login_password_input = (input("어드민 로그인 비밀번호: "))
    deal_admin_login_password_cell.value = deal_admin_login_password_input
    info_sheet['A2'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_password = deal_admin_login_password_input
    #print("1번째", deal_admin_login_password)

else:
    deal_admin_login_password = deal_admin_login_password_cell.value
    #print("2번째", deal_admin_login_password)

##### 셀러 계정 정보 #####
if deal_seller_login_id_cell.value is None:
    deal_seller_login_id_input = (input("셀러 로그인 ID: "))
    deal_seller_login_id_cell.value = deal_seller_login_id_input
    info_sheet['A3'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_id = deal_seller_login_id_input
    # print("1번째", deal_seller_login_id)
else:
    deal_seller_login_id = deal_seller_login_id_cell.value
    #print("2번째", deal_seller_login_id)


if deal_seller_login_password_cell.value is None:
    deal_seller_login_password_input = (input("셀러 로그인 비밀번호: "))
    deal_seller_login_password_cell.value = deal_seller_login_password_input
    info_sheet['A4'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_password = deal_seller_login_password_input
    #print("1번째", deal_seller_login_password)

else:
    deal_seller_login_password = deal_seller_login_password_cell.value
    #print("2번째", deal_seller_login_password)


##### WMS 계정 정보 #####
if wms_login_id_cell.value is None:
    wms_login_id_input = (input("WMS 로그인 ID: "))
    wms_login_id_cell.value = wms_login_id_input
    info_sheet['A5'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_id = wms_login_id_input
    # print("1번째", wms_login_id)
else:
    wms_login_id = wms_login_id_cell.value
    #print("2번째", wms_login_id)


if wms_login_passWord_cell.value is None:
    wms_login_passWord_input = (input("WMS 로그인 비밀번호: "))
    wms_login_passWord_cell.value = wms_login_passWord_input
    info_sheet['A6'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_passWord = wms_login_passWord_input
    #print("1번째", wms_login_passWord)

else:
    wms_login_passWord = wms_login_passWord_cell.value
    #print("2번째", wms_login_passWord)


#########################################################################################################################
# 구글 시트 연동

# 구글 json 경로
json_file_name_cell = info_sheet['A7']

if json_file_name_cell.value is None:
    json_file_name = filedialog.askopenfilename()    
    json_file_name_cell.value = json_file_name
    info_sheet['A7'].font = font_color
    info_workbook.save(info_file_path)
    # print("1번째", deal_admin_login_id)
else:
    json_file_name = json_file_name_cell.value
    #print("2번째", deal_admin_login_id)

# json_file_name = 'C:\\auto_json\\fulfillment-371610-dee41b117bdb.json'	# 구글 시트 jSON



scope = [
'https://spreadsheets.google.com/feeds',
'https://www.googleapis.com/auth/drive',
]

gc = gspread.service_account(filename=json_file_name)


# 구글 테스트 시나리오 엑셀 주소
google_url_cell = info_sheet['A8']

if google_url_cell.value is None:
    google_url_input = (input("구글 테스트 시나리오 엑셀 주소: "))
    google_url_cell.value = google_url_input
    info_sheet['A8'].font = font_color
    info_workbook.save(info_file_path)
    google_url = google_url_input
    #print("1번째", deal_seller_login_password)

else:
    google_url = google_url_cell.value
    #print("2번째", deal_seller_login_password)

# google_url = 'https://docs.google.com/spreadsheets/d/1fMf-pNUosGPMJ6evQJLTRRhAGpqZwEIhmC48wxLXHMQ/edit#gid=651282265' # 테스트 시나리오 엑셀 주소


# 구글 API 사용 이메일
google_email_cell = info_sheet['A9']

if google_email_cell.value is None:
    google_email_input = (input("구글 API 사용 이메일: "))
    google_email_cell.value = google_email_input
    info_sheet['A9'].font = font_color
    info_workbook.save(info_file_path)
    google_email = google_email_input
    #print("1번째", deal_seller_login_password)
else:
    google_email = google_email_cell.value

# google_email = 'client_email: fulfillment-test@fulfillment-371610.iam.gserviceaccount.com'

google_doc = gc.open_by_url(google_url)


# 구글 시트
google_sheet_cell = info_sheet['A10']

if google_sheet_cell.value is None:
    google_sheet_input = (input("구글 시트: "))
    google_sheet_cell.value = google_sheet_input
    info_sheet['A10'].font = font_color
    info_workbook.save(info_file_path)
    google_sheet = google_sheet_input
    #print("1번째", deal_seller_login_password)
else:
    google_sheet = google_sheet_cell.value

# google_sheet = '1.정상입고-사입앱' # 구글 시트


google_sheet = google_doc.worksheet(google_sheet)


# 출고 관련 정보(배송 요청) : 원본 파일
deal_out_excel_upload_path = r"c:\test\auto_out_org_file.xlsx" # Excel 파일 경로 설정

# 출고에 전달할 정보 : 원본 파일
deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정



#########################################################################################################################

#########################################################################################################################
# 크롭 탭 3개 실행
chrome_options = Options()
chrome_options.add_argument('--start-maximized')

driver = webdriver.Chrome(chrome_options=chrome_options)
driver.execute_script('window.open("about:blank", "_blank");')
driver.execute_script('window.open("about:blank", "_blank");')

tabs = driver.window_handles

driver.switch_to.window(tabs[0])
driver.get(deal_seller_url) #신상마켓 소매 -> 딜리버드 진입

driver.maximize_window()

driver.switch_to.window(tabs[1])
driver.get(deal_admin_url)

#driver.set_window_size(6000, 1024)
driver.switch_to.window(tabs[2])
driver.get(wms_url)

time.sleep(2)
# wms 각 항목 -> 조회 -> 리스트의 스크롤이 생기면서 데이터 로드의 어려움(현재 보여지는 화면의 데이터만 가져옴)
# pyautogui을 사용 -> wms 로그인 화면에서 글꼴 축소 -> 항목 조회 -> 리스트의 모든 데이터가 보이게 됨

action = ActionChains(driver)



#########################################################################################################################
##어드민 -> WMS 이동
#########################################################################################################################
driver.switch_to.window(tabs[2])


# 테스트를 위한 임시 저장 hihida
#deal_wms_purchase_number = "19674"

# wms 각 항목 -> 조회 -> 리스트의 스크롤이 생기면서 데이터 로드의 어려움(현재 보여지는 화면의 데이터만 가져옴)
# pyautogui을 사용 -> wms 로그인 화면에서 글꼴 축소 -> 항목 조회 -> 리스트의 모든 데이터가 보이게 됨
#pyautogui_count = 0
#while pyautogui_count < 8:
#    pyautogui.hotkey('ctrl', '-') # 화면 축소
#    pyautogui_count = pyautogui_count + 1

# time.sleep(130) # 사입 마감 처리 시간 후 로그인 시도 hihida



#WMS 로그인 진행
driver.find_element(By.ID, 'login').send_keys(wms_login_id)
driver.find_element(By.ID, 'password').send_keys(wms_login_passWord)
driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/form/div[2]/button').click()
time.sleep(2)
print("##########")
print("wms 로그인 완료")

google_sheet.update_acell('I25', 'OK')
print("I25 PASS")

## hihida 230102"""


#########################################################################################################################
##### 입고 관리 - 입고 진행현황 #####

# 입고 관리 -> 입고 진행현황 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 진행현황":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break


google_sheet.update_acell('I73', 'OK') 
print("73 PASS")
time.sleep(2)
print("##########")
print("입고 관리 - 입고 진행현황 이동")


# 230308리스트 상단에 선택되어 있는 칼럼들 삭제
# for문으로 돌려서 순차적으로 종료 하고 싶었지만, StaleElementReferenceException 오류 벌생
# 간단하게 첫번째를 여러번 돌리는 방법으로 긴급하게 수정
try:
    # 모든 ag-icon ag-icon-cancel 버튼을 찾음
    cancel_buttons = driver.find_elements(By.CSS_SELECTOR, "span.ag-icon.ag-icon-cancel")

    # 모든 ag-icon ag-icon-cancel 버튼을 클릭
    for cancel_button in cancel_buttons:
        cancel_button.click()
        break

    cancel_buttons = driver.find_elements(By.CSS_SELECTOR, "span.ag-icon.ag-icon-cancel")

    # 모든 ag-icon ag-icon-cancel 버튼을 클릭
    for cancel_button in cancel_buttons:
        cancel_button.click()
        break


    cancel_buttons = driver.find_elements(By.CSS_SELECTOR, "span.ag-icon.ag-icon-cancel")

    # 모든 ag-icon ag-icon-cancel 버튼을 클릭
    for cancel_button in cancel_buttons:
        cancel_button.click()
        break# 모든 ag-icon ag-icon-cancel 버튼을 클릭


    cancel_buttons = driver.find_elements(By.CSS_SELECTOR, "span.ag-icon.ag-icon-cancel")

    for cancel_button in cancel_buttons:
        cancel_button.click()
        break


    cancel_buttons = driver.find_elements(By.CSS_SELECTOR, "span.ag-icon.ag-icon-cancel")

    for cancel_button in cancel_buttons:
        cancel_button.click()
        break

except:
    pass



################################################
## 열(컬럼) 메뉴 - 특정 컬럼만 선택
# wms_test_result = driver.find_elements(By.CLASS_NAME, 'ag-side-button-button') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result[0].click() # 중간 오른쪽의 열 컬럼 버튼 클릭

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라짐
# wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-icon.ag-icon-columns') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result.click() # 중간 오른쪽의 열 컬럼 버튼 클릭
# time.sleep(3) # 0[열 컬럼] / 1[필터]

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라져 다시 수정
# <span class =ag-side-button-labe>에서 열(컬럼), 필터 중 열(컬럼) 찾아 버튼 클릭
time.sleep(2)
wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

time.sleep(3)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click() #전체 선택 버튼 클릭해 일단 전체 컬럼 선택
time.sleep(5)


#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click()#전체 선택 버튼 다시 클릭해 일단 전체 컬럼 선택 해제
#time.sleep(5)


# 체크박스 컬럼 선택
wms_test_result = driver.find_element(By.CSS_SELECTOR, "div>[aria-posinset='2']") # 체크박스 컬럼의 유일값을 찾음, aria-posinset='2'
time.sleep(5)


wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
time.sleep(5)

wms_test_result.find_element(By.ID, wms_test_result_chk).click()

#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 체크박스 유일값으로 동적으로 변화되는 aria-describedby ID값 취득
#time.sleep(5)


#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() 
time.sleep(5)





driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 주문번호 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("주문번호") # 컬럼 검색 필드 - 주문번호
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합


wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break


################################################



cell_data = google_sheet.acell('H74').value 

driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='도매 매장명']").send_keys(cell_data) # "도매 매장명"에 도매 명 입력
google_sheet.update_acell('I74', 'OK') 
print("74 PASS")
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='도매 매장명']").send_keys(Keys.ENTER) # 검색 적용
google_sheet.update_acell('I75', 'OK') 
print("75 PASS")

deal_wms_purchase_number = '20298'
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
google_sheet.update_acell('I78', 'OK') 
print("78 PASS")
time.sleep(2)



   
#리스트의 전체 체크 박스 선택
wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-header-cell.ag-header-cell-sortable.ag-focus-managed') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
wms_test_result_chk = wms_test_result.find_element(By.CLASS_NAME, 'ag-input-field-input.ag-checkbox-input')
wms_test_result_chk.click()



#wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result.find_element(By.ID, wms_test_result_chk).click()
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합
google_sheet.update_acell('I84', 'OK') 
print("84 PASS")

google_sheet.update_acell('I90', 'OK')
google_sheet.update_acell('I96', 'OK')
# hihida 230102"""
while(True):
    	pass
 