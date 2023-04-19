import selenium.webdriver.support.ui as ui

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog

from openpyxl.styles import Font

import pyautogui
import time
import sys
import json
import re
import openpyxl
import subprocess


import gspread
from oauth2client.service_account import ServiceAccountCredentials

import requests
from bs4 import BeautifulSoup



#########################################################################################################################
# 사입 요청 파일 다운로드
# https://docs.google.com/spreadsheets/d/19X1duCg7N2npHQHGu_pPcDaji_9pDWdI/edit#gid=1830395100
# 해당 엑셀 파일을 c:\test\ 에 저장



"""
##################### 중요!!!!!
# 테스트 전 복사해서 선언해야 함 deal_test_saip_excel_upload, buyer_wsIdx_name

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화11.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23163' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화12.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23164' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화13.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23165' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화14.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23166' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화15.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23167' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화16.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23168' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화17.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23169' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화18.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23170' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화19.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23171' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값

"""
# deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
# buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값


# 불러오기 창 생성
root = tk.Tk()
root.withdraw()
# excel_upload_file_path = filedialog.askopenfilename()
deal_test_saip_excel_upload = filedialog.askopenfilename()

# 선택된 파일 열기

# deal_test_saip_excel_upload = openpyxl.load_workbook(excel_upload_file_path)
# org_workbook = openpyxl.load_workbook(excel_upload_file_path)
# org_worsheet = org_workbook.active

# 사입 요청한 도매 wsIdx 값
buyer_wsIdx_name = input("Enter 도매 wsIndex number: ")


#########################################################################################################################
# 테스트 기본 설정
#########################################################################################################################
# 자동화 실행 시 기본 정보 로드 : ex)계정 정보 입력, 구글 접속 정보
deal_admin_login_id = ''
deal_admin_login_password = ''
deal_admin_url = 'https://dealibird.qa.sinsang.market/ssm_admins/sign_in'
deal_seller_login_id = ''
deal_seller_login_password = ''
deal_seller_url = 'https://vat.qa.sinsang.market/'

# WMS 테스트 기본 설정
wms_login_id = '' 									# WMS 로그인 ID
wms_login_passWord = ''  						# WMS 로그인 비번
wms_url = 'https://matrix-web.qa.sinsang.market/signin'

info_file_path = 'C:\\test\\info.xlsx'
try:
    info_workbook = openpyxl.load_workbook(info_file_path)
except FileNotFoundError:
    # 파일이 없을 경우 새로 생성
    info_workbook = openpyxl.Workbook()


if 'Sheet1' in info_workbook.sheetnames:
    info_sheet = info_workbook.active
else:
    info_sheet = info_workbook.create_sheet('Sheet1')


# info_file = askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "C:\\test\\info.xlsx")])

# info_file = askopenfilename(info_file_path)

# 엑셀 파일 열기
# info_workbook = openpyxl.load_workbook(info_file)

# 로그인 정보 입력 받기
#deal_admin_login_id = info_workbook.active['A1']
#deal_admin_login_password = info_workbook.active['B1']
deal_admin_login_id_cell = info_sheet['A1']
deal_admin_login_password_cell = info_sheet['A2']

deal_seller_login_id_cell = info_sheet['A3']
deal_seller_login_password_cell = info_sheet['A4']

wms_login_id_cell = info_sheet['A5']
wms_login_passWord_cell = info_sheet['A6']

font_color = Font(color='FFFFFF')
# print("0번째", deal_admin_login_id_cell)
# print("0번째", deal_admin_login_password_cell)

##### 어드민 계정 정보 #####
if deal_admin_login_id_cell.value is None:
    deal_admin_login_id_input = (input("어드민 로그인 ID: "))
    deal_admin_login_id_cell.value = deal_admin_login_id_input
    info_sheet['A1'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_id = deal_admin_login_id_input
    # print("1번째", deal_admin_login_id)
else:
    deal_admin_login_id = deal_admin_login_id_cell.value
    #print("2번째", deal_admin_login_id)


if deal_admin_login_password_cell.value is None:
    deal_admin_login_password_input = (input("어드민 로그인 비밀번호: "))
    deal_admin_login_password_cell.value = deal_admin_login_password_input
    info_sheet['A2'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_password = deal_admin_login_password_input
    #print("1번째", deal_admin_login_password)

else:
    deal_admin_login_password = deal_admin_login_password_cell.value
    #print("2번째", deal_admin_login_password)

##### 셀러 계정 정보 #####
if deal_seller_login_id_cell.value is None:
    deal_seller_login_id_input = (input("셀러 로그인 ID: "))
    deal_seller_login_id_cell.value = deal_seller_login_id_input
    info_sheet['A3'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_id = deal_seller_login_id_input
    # print("1번째", deal_seller_login_id)
else:
    deal_seller_login_id = deal_seller_login_id_cell.value
    #print("2번째", deal_seller_login_id)


if deal_seller_login_password_cell.value is None:
    deal_seller_login_password_input = (input("셀러 로그인 비밀번호: "))
    deal_seller_login_password_cell.value = deal_seller_login_password_input
    info_sheet['A4'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_password = deal_seller_login_password_input
    #print("1번째", deal_seller_login_password)

else:
    deal_seller_login_password = deal_seller_login_password_cell.value
    #print("2번째", deal_seller_login_password)


##### WMS 계정 정보 #####
if wms_login_id_cell.value is None:
    wms_login_id_input = (input("WMS 로그인 ID: "))
    wms_login_id_cell.value = wms_login_id_input
    info_sheet['A5'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_id = wms_login_id_input
    # print("1번째", wms_login_id)
else:
    wms_login_id = wms_login_id_cell.value
    #print("2번째", wms_login_id)


if wms_login_passWord_cell.value is None:
    wms_login_passWord_input = (input("WMS 로그인 비밀번호: "))
    wms_login_passWord_cell.value = wms_login_passWord_input
    info_sheet['A6'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_passWord = wms_login_passWord_input
    #print("1번째", wms_login_passWord)

else:
    wms_login_passWord = wms_login_passWord_cell.value
    #print("2번째", wms_login_passWord)


#########################################################################################################################
# 구글 시트 연동

# 구글 json 경로
json_file_name_cell = info_sheet['A7']

if json_file_name_cell.value is None:
    json_file_name = filedialog.askopenfilename()    
    json_file_name_cell.value = json_file_name
    info_sheet['A7'].font = font_color
    info_workbook.save(info_file_path)
    # print("1번째", deal_admin_login_id)
else:
    json_file_name = json_file_name_cell.value
    #print("2번째", deal_admin_login_id)

# json_file_name = 'C:\\auto_json\\fulfillment-371610-dee41b117bdb.json'	# 구글 시트 jSON



scope = [
'https://spreadsheets.google.com/feeds',
'https://www.googleapis.com/auth/drive',
]

gc = gspread.service_account(filename=json_file_name)


# 구글 테스트 시나리오 엑셀 주소
google_url_cell = info_sheet['A8']

if google_url_cell.value is None:
    google_url_input = (input("구글 테스트 시나리오 엑셀 주소: "))
    google_url_cell.value = google_url_input
    info_sheet['A8'].font = font_color
    info_workbook.save(info_file_path)
    google_url = google_url_input
    #print("1번째", deal_seller_login_password)

else:
    google_url = google_url_cell.value
    #print("2번째", deal_seller_login_password)

# google_url = 'https://docs.google.com/spreadsheets/d/1fMf-pNUosGPMJ6evQJLTRRhAGpqZwEIhmC48wxLXHMQ/edit#gid=651282265' # 테스트 시나리오 엑셀 주소


# 구글 API 사용 이메일
google_email_cell = info_sheet['A9']

if google_email_cell.value is None:
    google_email_input = (input("구글 API 사용 이메일: "))
    google_email_cell.value = google_email_input
    info_sheet['A9'].font = font_color
    info_workbook.save(info_file_path)
    google_email = google_email_input
    #print("1번째", deal_seller_login_password)
else:
    google_email = google_email_cell.value

# google_email = 'client_email: fulfillment-test@fulfillment-371610.iam.gserviceaccount.com'

google_doc = gc.open_by_url(google_url)


# 구글 시트
google_sheet_cell = info_sheet['A10']

if google_sheet_cell.value is None:
    google_sheet_input = (input("구글 시트: "))
    google_sheet_cell.value = google_sheet_input
    info_sheet['A10'].font = font_color
    info_workbook.save(info_file_path)
    google_sheet = google_sheet_input
    #print("1번째", deal_seller_login_password)
else:
    google_sheet = google_sheet_cell.value

# google_sheet = '1.정상입고-사입앱' # 구글 시트


google_sheet = google_doc.worksheet(google_sheet)


# 출고 관련 정보(배송 요청) : 원본 파일
deal_out_excel_upload_path = r"c:\test\auto_out_org_file.xlsx" # Excel 파일 경로 설정

# 출고에 전달할 정보 : 원본 파일
deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정



#########################################################################################################################
# 사입 요청 엑셀 파일 -> 사입앱 -> 사입 수량, 도매 단가, 최종 금액 변경에 대한 정보 입력

buyer_workbook = openpyxl.load_workbook(deal_test_saip_excel_upload)
buyer_worsheet = buyer_workbook.active

buyer_purchasedCount_01 = '' # 사입을 위한 모두 사입 수량
buyer_purchasedCount_02 = ''
buyer_purchasedCount_03 = ''

deal_prace_1 = '' # 사입 요청한 첫번째 SKU 도매 단가
deal_prace_2 = '' # 사입 요청한 두번째 SKU 도매 단가
deal_prace_3 = '' # 사입 요청한 세번째 SKU 도매 단가

    
buyer_purchasedCount_01 = buyer_worsheet['I4'].value
buyer_purchasedCount_02 = buyer_worsheet['I5'].value
buyer_purchasedCount_03 = buyer_worsheet['I6'].value

deal_prace_1 = buyer_worsheet['H4'].value
deal_prace_2 = buyer_worsheet['H5'].value
deal_prace_3 = buyer_worsheet['H6'].value

# # 사입을 위한 모두 사입 수량 01 -> 구글 시트에 업데이트
google_sheet.update_acell('H12', buyer_purchasedCount_01)
google_sheet.update_acell('H14', buyer_purchasedCount_02)
google_sheet.update_acell('H16', buyer_purchasedCount_03)

google_sheet.update_acell('H11', deal_prace_1)
google_sheet.update_acell('H13', deal_prace_2)
google_sheet.update_acell('H15', deal_prace_3)


#########################################################################################################################

#########################################################################################################################
# 크롭 탭 3개 실행
chrome_options = Options()
chrome_options.add_argument('--start-maximized')

driver = webdriver.Chrome(chrome_options=chrome_options)
driver.execute_script('window.open("about:blank", "_blank");')
driver.execute_script('window.open("about:blank", "_blank");')

tabs = driver.window_handles

driver.switch_to.window(tabs[0])
driver.get(deal_seller_url) #신상마켓 소매 -> 딜리버드 진입

driver.maximize_window()

driver.switch_to.window(tabs[1])
driver.get(deal_admin_url)

#driver.set_window_size(6000, 1024)
driver.switch_to.window(tabs[2])
driver.get(wms_url)

time.sleep(2)
# wms 각 항목 -> 조회 -> 리스트의 스크롤이 생기면서 데이터 로드의 어려움(현재 보여지는 화면의 데이터만 가져옴)
# pyautogui을 사용 -> wms 로그인 화면에서 글꼴 축소 -> 항목 조회 -> 리스트의 모든 데이터가 보이게 됨

action = ActionChains(driver)

# hihida 221230
#########################################################################################################################
########### 어드민 -> 딜리버드 셀러 이동 ##########
#########################################################################################################################
driver.switch_to.window(tabs[0])


#어드민 로그인 진행
#driver.find_element(By.ID, 'ssm_admin_email').send_keys(deal_admin_login_id)
#driver.find_element(By.ID, 'ssm_admin_password').send_keys(deal_admin_login_password)
#driver.find_element(By.NAME, 'commit').click()

#어드민에서 테스트 셀러 계정 검색 -> 딜리버드 파트너 센터 이동
#driver.find_element(By.ID, 'inline_search').send_keys(deal_seller_id)
#driver.find_element(By.ID, 'inline_search').send_keys(Keys.ENTER)
#element = driver.find_element(By.ID,'sel_date_1month')# 라디오 버튼 클릭
#driver.execute_script("arguments[0].click();", element)

#driver.implicitly_wait(5)
#셀러 ID 클릭
#time.sleep(5)
#driver.find_element(By.XPATH, '//*[@id="purchasesList"]/tbody/tr[1]/td[3]/a').click()


##신상마켓 소매 -> 딜리버드 이동
#신상마켓 로그인
time.sleep(3)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/header/div/div[2]/div[3]').click() # 로그인 버튼(페이지 상단)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[1]/input').send_keys(deal_seller_login_id) # 모달 / ID 입력
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[2]/input').send_keys(deal_seller_login_password) # 모달 / 비번 입력
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[3]/div[2]/div/button').click() # 모달 / 로그인 버튼
print("##########")
print("신상마켓 소매 로그인")



#딜리버드 바로가기 클릭
time.sleep(3)
driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[1]/div/span').click()
print("##########")
print("딜리버드 이동")


#########################################################################################################################
#### 딜리버드 -> 사입 요청 #####
time.sleep(5)

try:
    driver.find_element(By.XPATH,'//*[@id="navbarSupportedContent"]/ul/li[2]/a').click()
    time.sleep(5)
except:
    deal_link = driver.find_element(By.LINK_TEXT, ("사입요청"))
    deal_link.click()
    time.sleep(5)
    
print("##########")
print("사입 요청 시작")


# 기존에 등록했던 사입 요청 리스트가 있다면 입력 초기화 하고 새로 입력
deal_test_result = driver.find_element(By.XPATH,'//*[@id="purchase_totalCount"]') # 페이지 중간 왼쪽 -> 사입 요청 : X 값
deal_test_result_check = deal_test_result.text # 사입 요청 값에서 테스트 값을 저장

if deal_test_result_check != "-": # 사입 요청 건수가 0건일 경우 실행 되지 않음
    deal_table_count = int(deal_test_result_check) # 텍스트 값을 int형으로 형변환
    
    if deal_table_count > 0: # 사입 요청 건 수가 1건 이상 있을 경우
        driver.find_element(By.XPATH, '//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[1]').click() # 페이지 중간의 [입력초기화] 버튼
        time.sleep(3)
    
        driver.find_element(By.XPATH, '/html/body/div[5]/div/div[3]/button[3]').click() # 얼럿 -> 모두 초기화 ~ [예] 버튼
        time.sleep(2)

driver.find_element(By.XPATH, '//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[6]').click() # 페이지 중간의 [엑셀 업로드] 버튼
time.sleep(2)

#파일 업로드
up_load_file = driver.find_element(By.XPATH, '//*[@id="excel_file"]') # 모달 / 엑셀 업로드 양식 선태 -> 엑셀 파일 선택 Browse 버튼
up_load_file.send_keys(deal_test_saip_excel_upload)
driver.find_element(By.NAME, 'commit').click() # [저장] 버튼

print("##########")
print("파일 업로드 완료")





time.sleep(3)


#사입 요청 수량 체크
deal_test_result = driver.find_element(By.XPATH,'//*[@id="purchase_totalCount"]') # 페이지 중간 왼쪽 -> 사입 요청 : X 값
deal_test_result_check = deal_test_result.text

cell_data = google_sheet.acell('H8').value # 사입 요청 수량(리스트 수량) / 사입 요청 수량을 확인한다

if cell_data == deal_test_result_check:
    google_sheet.update_acell('I8', 'Pass') 
else:
    google_sheet.update_acell('I8', 'Failed')


#요청 가능 수량 체크
deal_test_result = driver.find_element(By.XPATH,'//*[@id="purchase_able_to_count"]') # 페이지 중간 왼쪽 -> 요청 가능 : X 값
deal_test_result_check = deal_test_result.text

cell_data = google_sheet.acell('H9').value # 요청 가능 수량(리스트 수량) / 요청 가능 수량을 확인한다.

if cell_data == deal_test_result_check:
    google_sheet.update_acell('I9', 'Pass') 
else:
    google_sheet.update_acell('I9', 'Failed')


#요청 불가능 수량 체크
deal_test_result = driver.find_element(By.XPATH,'//*[@id="purchase_unable_to_count"]') # 페이지 중간 왼쪽 -> 요청 불가능 : X 값
deal_test_result_check = deal_test_result.text

cell_data = google_sheet.acell('H10').value # 요청 불가능 수량(리스트 수량) / 요청 불가능 수량을 확인한다.

if cell_data == deal_test_result_check:
    google_sheet.update_acell('I10', 'Pass') 
else:
    google_sheet.update_acell('I10', 'Failed')
    
    


# 사입 요청 리스트(테이블) 체크
deal_table = driver.find_element(By.XPATH, '//*[@id="purchasesList"]') # 리스트(테이블) 전체 경로
deal_tbody = deal_table.find_element(By.TAG_NAME,'tbody')
deal_list_count = 0
deal_saib_do_store_name = "" # 도매 매장명


for tr in deal_tbody.find_elements(By.TAG_NAME,'tr'):
    for td in tr.find_elements(By.TAG_NAME,'td'):
        
        if deal_list_count == 11:
            cell_data = google_sheet.acell('H11').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I11', 'Pass')
            else:
                google_sheet.update_acell('I11', 'Failed')           
            
        if deal_list_count == 12:
            cell_data = google_sheet.acell('H12').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I12', 'Pass')
            else:
                google_sheet.update_acell('I12', 'Failed')           
        
        if deal_list_count == 13: # 테스트 데이터를 위해 엑셀에 업로드, 도매 매장명 # # 도매 명  -> 구글 시트에 업데이트
            deal_saib_do_store_name = td.get_attribute("innerText")
            # google_sheet.update_acell('H31', deal_saib_do_store_name)        
            # google_sheet.update_acell('H35', deal_saib_do_store_name)
            # google_sheet.update_acell('H39', deal_saib_do_store_name)
            # google_sheet.update_acell('H43', deal_saib_do_store_name)
            google_sheet.update_acell('H74', deal_saib_do_store_name)
            #google_sheet.update_acell('H100', deal_saib_do_store_name)
            #google_sheet.update_acell('H115', deal_saib_do_store_name)
            #google_sheet.update_acell('H127', deal_saib_do_store_name)        
            #google_sheet.update_acell('H131', deal_saib_do_store_name)        
            #google_sheet.update_acell('H135', deal_saib_do_store_name)        
                
        if deal_list_count == 37:
            cell_data = google_sheet.acell('H13').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I13', 'Pass')
            else:
                google_sheet.update_acell('I13', 'Failed')           
                
        if deal_list_count == 38:
            cell_data = google_sheet.acell('H14').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I14', 'Pass')
            else:
                google_sheet.update_acell('I14', 'Failed')           
                
        if deal_list_count == 63:
            cell_data = google_sheet.acell('H15').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I15', 'Pass')
            else:
                google_sheet.update_acell('I15', 'Failed')           
                
        if deal_list_count == 64:
            cell_data = google_sheet.acell('H16').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I16', 'Pass')
            else:
                google_sheet.update_acell('I16', 'Failed')           
   
        deal_list_count += 1


driver.find_element(By.XPATH, '//*[@id="purchasesList_wrapper"]/div[1]/div/div/button[9]').click() # 페이지 중간 오른쪽 [사입 요청하기] 버튼
time.sleep(2)


# 사입 요청 버튼
driver.find_element(By.XPATH, '/html/body/div[5]/div/div[3]/button[3]').click() # 얼럿 / X건의 상품을~~ 진행하시겠습니까? -> [네] 버튼
time.sleep(5)
print("##########")
print("사입 요청 완료")


# 결제 정보 모달
driver.find_element(By.ID, 'method_SINSANGPOINT').click() # 모달 / 결제 수단 선택 -> 신상캐시 버튼
time.sleep(2)
google_sheet.update_acell('I17', 'OK')
print(" 17 PASS")
print(" PASS")

try:
    driver.find_element(By.XPATH,'//*[@id="confirmCollapse"]/div[2]/div/label').click() # 모달 -> 이용 약관 -> 전체 동의합니다. 버튼
except:
    print("이용 약관 xpath 값 오류 except 진입")
    deal_test_result = driver.find_element(By.id, "policyAllCheck")
    deal_test_result.click()


try:
    driver.find_element(By.XPATH, '//*[@id="payment_button"]').click() # 모달 -> [결제하기] 버튼
except:
    print("모달 -> [결제하기] 버튼 xpath 값 오류 except 진입")
    deal_test_result = driver.find_element(By.id, "payment_button")
    deal_test_result.click()    
    

time.sleep(10)

print("##########")
print("결제 완료")
google_sheet.update_acell('I18', 'OK')
print("18 PASS")

#########################################################################################################################
# 결제 완료 후 사입 요청 페이지
# 중요 hihida 딜리버드 주문번호 저장 / 사입 요청 번호
deal_wms_purchase_number = driver.find_element(By.XPATH, '//*[@id="page-wrapper"]/div[2]/div[2]/div/div/div[1]/div[1]/h4/span')
deal_wms_purchase_number = deal_wms_purchase_number.text

# 딜리버드 주문번호 -> 구글 시트에 업데이트
google_sheet.update_acell('H43', deal_wms_purchase_number)
#google_sheet.update_acell('H50', deal_wms_purchase_number)
#google_sheet.update_acell('H54', deal_wms_purchase_number)
#google_sheet.update_acell('H78', deal_wms_purchase_number)
#google_sheet.update_acell('H102', deal_wms_purchase_number)
#google_sheet.update_acell('H118', deal_wms_purchase_number)


# 중요 hihida 딜리버드 상품코드 저장
deal_product_id_1 = driver.find_element(By.XPATH, '//*[@id="purchasesDetail"]/tbody/tr[1]/td[2]')
deal_product_id_2 = driver.find_element(By.XPATH, '//*[@id="purchasesDetail"]/tbody/tr[2]/td[2]')
deal_product_id_3 = driver.find_element(By.XPATH, '//*[@id="purchasesDetail"]/tbody/tr[3]/td[2]')

deal_product_id_1 = deal_product_id_1.text
deal_product_id_2 = deal_product_id_2.text
deal_product_id_3 = deal_product_id_3.text

# 딜리버드 상품코드  -> 구글 시트에 업데이트
#google_sheet.update_acell('H30', deal_product_id_1)
google_sheet.update_acell('H103', deal_product_id_1)
#google_sheet.update_acell('H119', deal_product_id_1)
#google_sheet.update_acell('H126', deal_product_id_1)

#google_sheet.update_acell('H34', deal_product_id_2)
google_sheet.update_acell('H106', deal_product_id_2)
#google_sheet.update_acell('H121', deal_product_id_2)
#google_sheet.update_acell('H130', deal_product_id_2)

#google_sheet.update_acell('H38', deal_product_id_3)
google_sheet.update_acell('H109', deal_product_id_3)
#google_sheet.update_acell('H123', deal_product_id_3)
#google_sheet.update_acell('H134', deal_product_id_3)




#########################################################################################################################
########## 어드민 사입마감 처리 ##########
#########################################################################################################################

driver.switch_to.window(tabs[1])


#어드민 로그인 진행
time.sleep(10)
driver.find_element(By.ID, 'ssm_admin_email').send_keys(deal_admin_login_id)
driver.find_element(By.ID, 'ssm_admin_password').send_keys(deal_admin_login_password)
driver.find_element(By.NAME, 'commit').click()
time.sleep(2)
print("##########")
print("어드민 로그인 완료")

deal_admin_saip_end_url = 'https://dealibird.qa.sinsang.market/admin/pps/purchase_schedules/manual_transfer/' + deal_wms_purchase_number
# admin_response = requests.post(url=deal_admin_saip_end_url) # 사입 마감 처리
driver.get(deal_admin_saip_end_url) # 사입 마감 처리



print("##########")
# print("사입 마감 처리 완료", admin_response)
print("사입 마감 처리 완료")
# hihida 221230
google_sheet.update_acell('I19', 'OK')
print("19 PASS")
time.sleep(10)


#########################################################################################################################
########### 사입앱 로그인 ##########
#########################################################################################################################
buyer_login_url = 'https://buyer.qa.sinsang.market/api/v1/session' # 사입 로그인 URL
buyer_login_header = {'Content-Type' : 'application/json', "User-Agent" : "Mozilla/5.0"} # 로그인 시 헤더
buyer_login_data = {
    'password':'1234',
    'user':'qa_smkim'
} # 로그인 시 Body 정보 : 로그인 계정

buyer_response = requests.post(url=buyer_login_url, headers=buyer_login_header, params=buyer_login_data) # 로그인 시도
print("사입앱 로그인 성공\n", buyer_response)

buyer_login_content = buyer_response.content # 로그인 후 리턴되는 값(여러 정보가 있음)
buyer_login_content_data = json.loads(buyer_login_content) # JSON 문자열을 Python 객체로 변환
buyer_login_content_accesstoken = buyer_login_content_data["content"]["accessToken"] # accessToken 저장




#########################################################################################################################
########### 사입 리스트 상세 : 사입 예정 SKU의 ID 가져오기 ##########

buyer_login_accesstoken_header = {'Content-Type': 'application/json',
#'access-token': login_accesstoken,
'access-token': buyer_login_content_accesstoken,
'User-Agent': 'Mozilla/5.0',
'Cache-Control': 'no-cache',
'Accept': '*/*',
'Host': 'buyer.qa.sinsang.market',
'Accept-Encoding': 'gzip, deflate, br',
'Connection': 'keep-alive'} # accessToken을 가지고 사입 리스트 상세 조회


# id_search_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_detail?wsIdx=23124&orderType=purchase&warehouse=B1'

# buyer_id_search_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_detail?wsIdx='+ buyer_wsIdx_name + '&orderType=purchase&warehouse=B1' # 사입 리스트 상세 조회 URL
buyer_id_search_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_detail?wsIdx='+ buyer_wsIdx_name + '&orderType=purchase' # 사입 리스트 상세 조회 URL
print(buyer_wsIdx_name)

buyer_response = requests.get(url=buyer_id_search_url, headers=buyer_login_accesstoken_header) # 사입 리스트 상세 조회 시도
print("사입 리스트 상세 조회 성공\n", buyer_response)

buyer_id_search_content = buyer_response.content# 조회 후 리턴되는 값(여러 정보가 있음)

buyer_id_search_content_data = json.loads(buyer_id_search_content) # JSON 문자열을 Python 객체로 변환

buyer_id_search_content_ID_data = [] # 여러개의 ID 정보 저장을 위한 배열
buyer_id_search_int = int(0) # 배열 Len 체크
for product in buyer_id_search_content_data["content"]["products"]:
    buyer_id_search_content_ID_data.append(product["id"]) # id 정보를 배열(id_search_content_ID_data)에 저장
    print(buyer_id_search_int, "번째 ID는", buyer_id_search_content_ID_data[buyer_id_search_int])
    buyer_id_search_int = buyer_id_search_int +1

buyer_id_search_int = buyer_id_search_int -1 # 최종 배열 길이 체크


#########################################################################################################################
########### 사입 상품 옵션 저장 : SKU의 ID로 모두 사입으로 전송 ##########

buyer_order_status_url = 'https://buyer.qa.sinsang.market/buyer/api/dealibird/buying/order_status' # 사입 상태 전송 URL
                    
buyer_order_status_data = {
    "orderType" : "purchase", # 신규 주문
	"items": [
		{
            "id": buyer_id_search_content_ID_data[0],
            # "purchasedCount" : 5
            "purchasedCount" : buyer_purchasedCount_01
		},
        {
            "id": buyer_id_search_content_ID_data[1],
            "purchasedCount" : buyer_purchasedCount_02
            # "purchasedCount" : 8
		},
        {
            "id": buyer_id_search_content_ID_data[2],
            "purchasedCount" : buyer_purchasedCount_03
            # "purchasedCount" : 10
		}
	]
} # 로그인 시 Body 정보 : 로그인 계정


buyer_response = requests.post(url=buyer_order_status_url, headers=buyer_login_accesstoken_header, data=json.dumps(buyer_order_status_data)) # 사입 상태 전송 시도

print("사입앱 사입 성공 전달 성공\n", buyer_response)

google_sheet.update_acell('I16', 'OK')
print("16 PASS")


#########################################################################################################################
##어드민 -> WMS 이동
#########################################################################################################################
driver.switch_to.window(tabs[2])


# 테스트를 위한 임시 저장 hihida
#deal_wms_purchase_number = "19674"

# wms 각 항목 -> 조회 -> 리스트의 스크롤이 생기면서 데이터 로드의 어려움(현재 보여지는 화면의 데이터만 가져옴)
# pyautogui을 사용 -> wms 로그인 화면에서 글꼴 축소 -> 항목 조회 -> 리스트의 모든 데이터가 보이게 됨
#pyautogui_count = 0
#while pyautogui_count < 8:
#    pyautogui.hotkey('ctrl', '-') # 화면 축소
#    pyautogui_count = pyautogui_count + 1

# time.sleep(130) # 사입 마감 처리 시간 후 로그인 시도 hihida
time.sleep(10) # 테스트 임시


#WMS 로그인 진행
driver.find_element(By.ID, 'login').send_keys(wms_login_id)
driver.find_element(By.ID, 'password').send_keys(wms_login_passWord)
driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/form/div[2]/button').click()
time.sleep(2)
print("##########")
print("wms 로그인 완료")

google_sheet.update_acell('I25', 'OK')
print("I25 PASS")

## hihida 230102"""
#########################################################################################################################
##### 입고 관리 - 입고 대기 #####
time.sleep(2)
# 입고 관리 -> 입고 대기 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 대기":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break   
    
 

print("##########")
print("입고 관리 - 입고 대기 이동")
google_sheet.update_acell('I42', 'OK')
print("I42 PASS")
time.sleep(2)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
time.sleep(2)
google_sheet.update_acell('I43', 'OK') 
print("I43 PASS")


# 총 주문 수 갯수 가져오기 old
#try:
#    wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[3]/div/div/div/h2[1]')
#except:
#    wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div[2]/div[2]/div/div/div/h2[1]')
# wms_test_result_check = wms_test_result.text
# wms_test_result_check = wms_test_result_check.split(" ") # 공백만 제거 하고 배열에 입력
# wms_test_result_check = wms_test_result_check[3]


# 총 주문 수 갯수 가져오기(230116)
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'h2.MuiTypography-root')
for wms_str_loop in wms_test_result:
    #print("for 문",wms_str_loop.text,"\n")
    if "총 주문 수" in wms_str_loop.text:
        #print("가져오기 시도\n")
        wms_str_loop_result = wms_str_loop.text
        #print(wms_str_loop_result, "가져오기 완료\n")
        break
wms_str_loop_result = wms_str_loop_result.replace(' ', '') # 공백만 제거 하고 배열에 입력
wms_str_loop_result = wms_str_loop_result.replace("총주문수", '') # 나머지 내용 제거
wms_test_result_check = wms_str_loop_result.replace('개', '') # 나머지 내용 제거

print("총 주문 수 갯수", wms_test_result_check)
cell_data = google_sheet.acell('H44').value # 총 주문수를 확인한다.

if cell_data == wms_test_result_check:
    google_sheet.update_acell('I44', 'Pass')
else:
    google_sheet.update_acell('I44', 'Failed')

print("##########")
print("입고 대기 - 총 주문 수 갯수 가져오기")


#총 상품수(sku) 갯수 가져오기 old
#try:
#    wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[3]/div/div/div/h2[2]')
#except:
#    wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div[2]/div[2]/div/div/div/h2[2]')
#wms_test_result_check = wms_test_result.text
#wms_test_result_check = wms_test_result_check.split(" ") # 공백만 제거 하고 배열에 입력
#wms_test_result_check = wms_test_result_check[2]
#wms_search_list_row = wms_test_result_check # 하단에서 리스트의 ROW수 계산을 위한 데이터 입력

# 총 상품수(sku) 갯수 가져오기(230116)
time.sleep(3)
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'h2.MuiTypography-root')
for wms_str_loop in wms_test_result:
    #print("for 문",wms_str_loop.text,"\n")
    if "총 상품수(sku)" in wms_str_loop.text:
       #print("가져오기 시도\n")
        wms_str_loop_result = wms_str_loop.text
        #print(wms_str_loop_result, "가져오기 완료\n")
        break

wms_str_loop_result = wms_str_loop_result.replace(' ', '') # 공백만 제거 하고 배열에 입력
wms_str_loop_result = wms_str_loop_result.replace("총상품수(sku)", '') # 나머지 내용 제거
wms_test_result_check = wms_str_loop_result.replace('개', '') # 나머지 내용 제거
print("총 상품수(sku)", wms_test_result_check)

cell_data = google_sheet.acell('H45').value # 총 상품 수를 확인한다

if cell_data == wms_test_result_check:
    google_sheet.update_acell('I45', 'Pass')
else:
    google_sheet.update_acell('I45', 'Failed')

print("##########")
print("입고 대기 - 총 상품수(sku) 갯수 가져오기")



#총 상품수량 갯수 가져오기 old
#try:
#    wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[3]/div/div/div/h2[3]')
#except:
#    wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div[2]/div[2]/div/div/div/h2[3]')
#wms_test_result_check = wms_test_result.text
#wms_test_result_check = wms_test_result_check.split(" ") # 공백만 제거 하고 배열에 입력
#wms_test_result_check = wms_test_result_check[2]


#총 상품수량 갯수 가져오기 (230116)
time.sleep(3)
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'h2.MuiTypography-root')

for wms_str_loop in wms_test_result:
    #print("for 문",wms_str_loop.text,"\n")
    if "총 상품수량" in wms_str_loop.text:
       #print("가져오기 시도\n")
        wms_str_loop_result = wms_str_loop.text
        #print(wms_str_loop_result, "가져오기 완료\n")
        break

wms_str_loop_result = wms_str_loop_result.replace(' ', '') # 공백만 제거 하고 배열에 입력
wms_str_loop_result = wms_str_loop_result.replace("총상품수량", '') # 나머지 내용 제거
wms_test_result_check = wms_str_loop_result.replace('개', '') # 나머지 내용 제거

cell_data = google_sheet.acell('H46').value # 총 상품 수량을 확인한다.

if cell_data == wms_test_result_check:
    google_sheet.update_acell('I46', 'Pass')
else:
    google_sheet.update_acell('I46', 'Failed')

print("##########")
print("입고 대기 - 총 상품수량 갯수 가져오기")


# 입고 요청 데이터 확인(테이블 - 리스트)
#wms_search_list = driver.find_element(By.CSS_SELECTOR,'div[class="ag-pinned-left-cols-container"]') # 총 SKU 수 확인, 각 행마다 데이터 검증을 위한 row 확인
print("##########")
print("입고 대기 - 입고 요청 데이터 확인(테이블 - 리스트) 시작")


wms_search_list_row_result = driver.find_element(By.CSS_SELECTOR,'div[class="ag-center-cols-container"]')
wms_search_list_row_id_result = wms_search_list_row_result.find_element(By.CSS_SELECTOR, 'div[row-index="0"]') # 리스트 row 마다 ID가져오기
#wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="4"]') # 소봉바코드 : 04
wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="purchase_barcode"]') # 소봉바코드
#wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="purchaseBarcode"]') # 소봉바코드
wms_small_bag_barcode = wms_search_list_row_result_check_1.text # 소봉 바코드 저장
google_sheet.update_acell('H49', wms_small_bag_barcode)
print("입고 대기 - 소봉바코드   ", wms_small_bag_barcode)

google_sheet.update_acell('I47', 'OK') 
print("47 PASS")



#########################################################################################################################
##### 입고 관리 - 입고 검수진행 #####
time.sleep(2)
# 입고 관리 -> 입고 검수진행 이동 old
# driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[1]/div/a[3]').click() # 입고 관리 -> 입고 검수진행 이동

# 입고 관리 -> 입고 검수진행 이동(230116)
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 검수진행":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break

time.sleep(2)
print("##########")
print("입고 관리 - 입고 검수진행 이동")
google_sheet.update_acell('I48', 'OK') 
print("48 PASS")



time.sleep(5)
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='바코드(도매처 소봉, 상품)를 스캔해 주세요']").send_keys(wms_small_bag_barcode) # "바코드 입력"에 입고대기에서 복사한 바코드 입력
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='바코드(도매처 소봉, 상품)를 스캔해 주세요']").send_keys(Keys.ENTER) # 검색 적용
google_sheet.update_acell('I49', 'OK') 
print("49 PASS")





## hihida 230102"""
## 열(컬럼) 메뉴 - 특정 컬럼만 선택
# wms_test_result = driver.find_elements(By.CLASS_NAME, 'ag-side-button-button') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result[0].click() # 중간 오른쪽의 열 컬럼 버튼 클릭

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라짐
# wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-icon.ag-icon-columns') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result.click() # 중간 오른쪽의 열 컬럼 버튼 클릭
# time.sleep(3) # 0[열 컬럼] / 1[필터]

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라져 다시 수정
# <span class =ag-side-button-labe>에서 열(컬럼), 필터 중 열(컬럼) 찾아 버튼 클릭
time.sleep(2)
wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

time.sleep(3)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click() #전체 선택 버튼 클릭해 일단 전체 컬럼 선택
time.sleep(3)

driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click() #전체 선택 버튼 클릭해 일단 전체 컬럼 해제
time.sleep(3)

# 주문번호 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("주문번호") # 컬럼 검색 필드 - 주문번호 입력
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(3)


#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 주문수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("주문수량") # 컬럼 검색 필드 - 주문수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)


# 장끼수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("장끼수량") # 컬럼 검색 필드 - 장끼수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 장끼수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)


# 낱개수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("낱개수량") # 컬럼 검색 필드 - 낱개수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 낱개수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 센터입고수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("센터입고수량") # 컬럼 검색 필드 - 센터입고수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 센터입고수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 입고상태 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("입고상태") # 컬럼 검색 필드 - 입고상태
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 입고상태 체크박스 xPATH값 조합
#time.sleep(5)

wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='셀러ID 필터 입력']").send_keys(deal_seller_login_id) # 테이블(리스트) -> 셀러ID 입력
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
time.sleep(2)
google_sheet.update_acell('I50', 'OK') 
print("50 PASS")

# 신규 주문 수량 확인
wms_test_result = driver.find_element(By.CLASS_NAME, 'MuiButtonBase-root.MuiTab-root.MuiTab-textColorPrimary.Mui-selected') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
wms_test_result_check = wms_test_result.text
wms_test_result_check = wms_test_result_check.replace('신규 주문 (','')
wms_test_result_check = wms_test_result_check.replace('건)','') # 텍스트에서 수량만 취득하기 위해 나머지 텍스트 삭제


# wms_test_result = driver.find_element(By.CLASS_NAME, 'MuiButtonBase-root.MuiTab-root.MuiTab-textColorPrimary.Mui-selected.css-1fs0d0o') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result = wms_test_result.get_attribute("ID") # 동적으로 변화되는 ID값 취득
# wms_search_list_row_index_str = "//*[@id=\"" + wms_test_result + "\"]/span[1]" # 버튼[(신규 주문 (X건)]의 xPATH값 조합
# wms_test_result = driver.find_element(By.XPATH, wms_search_list_row_index_str) # 버튼[(신규 주문 (X건)] -> 신규 주문 (X건) 값 취득
# wms_test_result_check = wms_test_result.text
# wms_test_result_check = wms_test_result_check.replace('신규 주문 (','')
# wms_test_result_check = wms_test_result_check.replace('건)','') # 텍스트에서 수량만 취득하기 위해 나머지 텍스트 삭제


cell_data = google_sheet.acell('H51').value # 신규 주문 수량 확인한다.

if cell_data == wms_test_result_check:
    google_sheet.update_acell('I51', 'Pass')
else:
    google_sheet.update_acell('I51', 'Failed')

print("##########")
print("입고 검수진행 - 신규 주문 수량 확인")



#바코드 출력 후 총 도매단가, 총 장끼 수량 가져오기
#wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[5]/h2[2]') # 총 도매단가 / 위치 강제
#wms_test_result_check = wms_test_result.text
time.sleep(15) # 총 도매 단가를 가져오려면 일정 시간 대기를 해야 로드 할 수 있음
# 총 도매단가 가져오기(230116)
cell_data = google_sheet.acell('H52').value # 총 도매단가 확인한다.
print("입고 검수진행 - cell_data!", cell_data,("!"))
try:
    wms_test_result_check = driver.find_element(By.CLASS_NAME, 'MuiTypography-root.MuiTypography-h2.css-i8nswv').text
except:
    # (230116)
    print("try try try try try try")
    time.sleep(3)
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'h2.MuiTypography-root')
    for wms_str_loop in wms_test_result:
        #print("for 문",wms_str_loop.text,"\n")
        if cell_data in wms_str_loop.text:
            #print("가져오기 시도\n")
            wms_test_result_check = wms_str_loop.text
            #print(wms_str_loop_result, "가져오기 완료\n")
            break
    
    
print("입고 검수진행 - wms_test_result_check!", wms_test_result_check,("!"))

time.sleep(5)

wms_test_result_check = wms_test_result_check.replace(' ','')
print("입고 검수진행 - wms_test_result_check.replace!", wms_test_result_check,("!"))
cell_data = cell_data.replace(' ','')
print("입고 검수진행 - cell_data.replace!", cell_data,("!"))


if cell_data == wms_test_result_check: # hihida 확인해야 함 45,400  @@@@  / 45,400 @@@@  -> 각 값의 뒷 공백 길이가 다름, 그래서 failed
    google_sheet.update_acell('I52', 'Pass')
else:
    google_sheet.update_acell('I52', 'Failed')

time.sleep(20)
#wms_test_result = driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[5]/h2[4]') # 총 장끼 수량 / 위치 강제
#wms_test_result_check = wms_test_result.text
# 총 장끼수량 가져오기 (230116)
cell_data = google_sheet.acell('H53').value # 총 장끼수량 확인한다.
try:
    wms_test_result_check = driver.find_element(By.CLASS_NAME, 'MuiTypography-root.MuiTypography-h2.css-1tq6ygv').text
except:
    # (230116)
    time.sleep(3)
    print("try try try try try try")
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'h2.MuiTypography-root')
    for wms_str_loop in wms_test_result:
        #print("for 문",wms_str_loop.text,"\n")
        if cell_data in wms_str_loop.text:
            #print("가져오기 시도\n")
            wms_str_loop = wms_str_loop.text
            #print(wms_str_loop_result, "가져오기 완료\n")
            wms_test_result_check = wms_str_loop_result.replace('개', '') # 나머지 내용 제거
            break
   


if cell_data == wms_test_result_check:
    google_sheet.update_acell('I53', 'Pass')
else:
    google_sheet.update_acell('I53', 'Failed')


print("##########")
print("입고 검수진행 - 바코드 출력 후 총 도매단가, 총 장끼 수량 가져오기 종료")





# 입고 요청 데이터 확인(테이블 - 리스트)
#wms_search_list = driver.find_element(By.CSS_SELECTOR,'div[class="ag-pinned-left-cols-container"]') # 총 SKU 수 확인, 각 행마다 데이터 검증을 위한 row 확인
print("##########")
print("입고 검수진행 - 입고 요청 데이터 확인(테이블 - 리스트) 시작")



# wms_search_list_row_index_ini = 3 
wms_search_list_row_index_ini = 100000
wms_search_list_row = int(0)

wms_search_list_row_index_ini = int(wms_search_list_row_index_ini) # 반복문을 실행하기 위한 int -> 형변환
wms_search_list_row_index_ini = wms_search_list_row_index_ini - 1  # 리스트의 배열이 0부터 시작하기 떄문에 -1

wms_search_list_row_result = driver.find_element(By.CSS_SELECTOR,'div[class="ag-center-cols-container"]')

try:
    while wms_search_list_row_index_ini >= wms_search_list_row : # 리스트 row 수 만큼 실행
        wms_search_list_row = str(wms_search_list_row) # row 0 부터 wms_search_list.find_element에 입력하기 위해 str -> 형변환
        wms_search_list_row_index_str = "div[row-index=\"" + wms_search_list_row + "\"]" # wms_search_list.find_element 에 row 0부터 조회를 위한 값 합치기
        wms_search_list_row_id_result = wms_search_list_row_result.find_element(By.CSS_SELECTOR, wms_search_list_row_index_str) # 리스트 row 마다 ID가져오기
        
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="12"]') # 합계(주문수량) : 12
        wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="buy_stock"]') # 합계(주문수량)
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="buyStock"]') # 합계(주문수량)
        wms_search_list_row_result_check_1 = wms_search_list_row_result_check_1.text
        print("입고 검수진행 - wms_search_list_row_result_check_1   ", wms_search_list_row_result_check_1)
        
        
        #wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="14"]') # 합계(장끼수량) : 14 -> 하위 버튼값 확인
        wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="jangkki_stock"]') # 합계(장끼수량)
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="jangkkiStock"]') # 합계(장끼수량)
        wms_search_list_row_result_check_2 = wms_search_list_row_result_check_2.text
        print("입고 검수진행 - wms_search_list_row_result_check_2   ", wms_search_list_row_result_check_2)
        
        
        #wms_search_list_row_result_check_3 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="15"]') # 합계(낱개수량) : 15
        wms_search_list_row_result_check_3 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="piece_stock"]') # 합계(낱개수량)
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="pieceStock"]') # 합계(낱개수량)
        wms_search_list_row_result_check_3 = wms_search_list_row_result_check_3.text
        print("입고 검수진행 - wms_search_list_row_result_check_3   ", wms_search_list_row_result_check_3)
        
        
        #wms_search_list_row_result_check_4 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="16"]') # 합계(센터입고수량) : 16 -> 하위 버튼값 확인
        wms_search_list_row_result_check_4 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="in_stock"]') # 합계(센터입고수량)
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="inStock"]') # 합계(센터입고수량)
        wms_search_list_row_result_check_4 = wms_search_list_row_result_check_4.text
        print("입고 검수진행 - wms_search_list_row_result_check_4   ", wms_search_list_row_result_check_4)
        
        
        #wms_search_list_row_result_check_5 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="19"]') # 입고상태  : 19
        wms_search_list_row_result_check_5 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="status_name"]') # 입고상태
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="statusName"]') # 입고상태
        wms_search_list_row_result_check_5 = wms_search_list_row_result_check_5.text
        print("입고 검수진행 - wms_search_list_row_result_check_5   ", wms_search_list_row_result_check_5)


        print("##########")
        print("입고 검수진행 - 입고 요청 데이터 확인(테이블 - 리스트) 종료")
    

        wms_search_list_row = int(wms_search_list_row) # while을 실행하기 위해 int -> 형변환
        
        if wms_search_list_row == 0 :
            cell_data = google_sheet.acell('H55').value # 도매 상품명_자동화001 -> 합계(사입성공수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I55', 'Pass')
            else:
                google_sheet.update_acell('I55', 'Failed')
    
            
            cell_data = google_sheet.acell('H56').value # 도매 상품명_자동화001 -> 합계(장끼수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I56', 'Pass')
            else:
                google_sheet.update_acell('I56', 'Failed')



            cell_data = google_sheet.acell('H57').value # 도매 상품명_자동화001 -> 합계(낱개수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_3:
                google_sheet.update_acell('I57', 'Pass')
            else:
                google_sheet.update_acell('I57', 'Failed')
    
    
            cell_data = google_sheet.acell('H58').value # 도매 상품명_자동화001 -> 합계(센터입고수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_4:
                google_sheet.update_acell('I58', 'Pass')
            else:
                google_sheet.update_acell('I58', 'Failed')


            cell_data = google_sheet.acell('H59').value # 도매 상품명_자동화001 -> 입고상태 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_5:
                google_sheet.update_acell('I59', 'Pass')
            else:
                google_sheet.update_acell('I59', 'Failed')


    
    
        if wms_search_list_row == 1 :
            cell_data = google_sheet.acell('H60').value # 도매 상품명_자동화002 -> 합계(사입성공수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I60', 'Pass')
            else:
                google_sheet.update_acell('I60', 'Failed')
    
            
            cell_data = google_sheet.acell('H61').value # 도매 상품명_자동화002 -> 합계(장끼수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I61', 'Pass')
            else:
                google_sheet.update_acell('I61', 'Failed')



            cell_data = google_sheet.acell('H62').value # 도매 상품명_자동화002 -> 합계(낱개수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_3:
                google_sheet.update_acell('I62', 'Pass')
            else:
                google_sheet.update_acell('I62', 'Failed')
    
    
            cell_data = google_sheet.acell('H63').value # 도매 상품명_자동화002 -> 합계(센터입고수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_4:
                google_sheet.update_acell('I63', 'Pass')
            else:
                google_sheet.update_acell('I63', 'Failed')


            cell_data = google_sheet.acell('H64').value # 도매 상품명_자동화002 -> 입고상태 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_5:
                google_sheet.update_acell('I64', 'Pass')
            else:
                google_sheet.update_acell('I64', 'Failed')

    
        if wms_search_list_row == 2 :
            cell_data = google_sheet.acell('H65').value # 도매 상품명_자동화003 -> 합계(사입성공수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I65', 'Pass')
            else:
                google_sheet.update_acell('I65', 'Failed')
    
            
            cell_data = google_sheet.acell('H66').value # 도매 상품명_자동화003 -> 합계(장끼수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I66', 'Pass')
            else:
                google_sheet.update_acell('I66', 'Failed')



            cell_data = google_sheet.acell('H67').value # 도매 상품명_자동화003 -> 합계(낱개수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_3:
                google_sheet.update_acell('I67', 'Pass')
            else:
                google_sheet.update_acell('I67', 'Failed')
    
    
            cell_data = google_sheet.acell('H68').value # 도매 상품명_자동화003 -> 합계(센터입고수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_4:
                google_sheet.update_acell('I68', 'Pass')
            else:
                google_sheet.update_acell('I68', 'Failed')


            cell_data = google_sheet.acell('H69').value # 도매 상품명_자동화003 -> 입고상태 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_5:
                google_sheet.update_acell('I69', 'Pass')
            else:
                google_sheet.update_acell('I69', 'Failed')

    
        wms_search_list_row = wms_search_list_row + 1  # row 증가
except:
    pass


# 전체 바코드 출력하기
time.sleep(10)
print("##########")
print("입고 검수진행 - 전체 바코드 출력) ] 버튼 시작") # [전체 바코드 출력]버튼 클릭한다.

try:
    driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/button').click() # 중간 -> [전체 바코드 출력] 버튼 클릭
except:
    print("try try try try try try")
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeLarge.MuiButton-containedSizeLarge.MuiButtonBase-root')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "전체 바코드 출력":
            #print("클릭 시도\n")
            wms_str_loop.click()
            #print("클릭 완료\n")
            break

google_sheet.update_acell('I70', 'OK') 
print("70 PASS")
time.sleep(10)


#driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/div[2]/div/div/button').click() # 모달 -> 전체 상품 바코드 출력 -> 진행 클릭
driver.find_element(By.XPATH,'/html/body/div[3]/div[3]/div[2]/button[2]').click() # 모달 -> 전체 상품 바코드 출력 -> 진행 클릭
google_sheet.update_acell('I71', 'OK') 
print("71 PASS")
time.sleep(10)

# 젙체 바코드 출력 후 시리얼 저장
alert = driver.switch_to.alert 
alert_barcode_all_print_text = alert.text


alert.accept() # 얼럿 확인
print("##########")
print("입고 검수진행 - 전체 바코드 출력) ] 버튼 종료")
google_sheet.update_acell('I72', 'OK') 
print("72 PASS")
time.sleep(5)



#########################################################################################################################
##### 입고 관리 - 입고 진행현황 #####

# 입고 관리 -> 입고 진행현황 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 진행현황":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break


google_sheet.update_acell('I73', 'OK') 
print("73 PASS")
time.sleep(2)
print("##########")
print("입고 관리 - 입고 진행현황 이동")

################################################
## 열(컬럼) 메뉴 - 특정 컬럼만 선택
# wms_test_result = driver.find_elements(By.CLASS_NAME, 'ag-side-button-button') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result[0].click() # 중간 오른쪽의 열 컬럼 버튼 클릭

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라짐
# wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-icon.ag-icon-columns') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result.click() # 중간 오른쪽의 열 컬럼 버튼 클릭
# time.sleep(3) # 0[열 컬럼] / 1[필터]

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라져 다시 수정
# <span class =ag-side-button-labe>에서 열(컬럼), 필터 중 열(컬럼) 찾아 버튼 클릭
time.sleep(2)
wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

time.sleep(3)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click() #전체 선택 버튼 클릭해 일단 전체 컬럼 선택
time.sleep(5)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click()#전체 선택 버튼 다시 클릭해 일단 전체 컬럼 선택 해제
time.sleep(5)


# 체크박스 컬럼 선택
wms_test_result = driver.find_element(By.CSS_SELECTOR, "div>[aria-posinset='2']") # 체크박스 컬럼의 유일값을 찾음, aria-posinset='2'
time.sleep(5)


wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
time.sleep(5)

wms_test_result.find_element(By.ID, wms_test_result_chk).click()

#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 체크박스 유일값으로 동적으로 변화되는 aria-describedby ID값 취득
#time.sleep(5)


#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() 
time.sleep(5)



# 사입 성공 수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("사입성공수량") # 컬럼 검색 필드 - 사입 성공 수량 입력
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)


#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 사입성공수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 장끼수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("장끼수량") # 컬럼 검색 필드 - 장끼수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 장끼수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 낱개수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("낱개수량") # 컬럼 검색 필드 - 낱개수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 낱개수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 센터입고수량 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("센터입고수량") # 컬럼 검색 필드 - 장끼수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 센터입고수량 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 입고상태 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("입고상태") # 컬럼 검색 필드 - 장끼수량
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 입고상태 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 주문번호 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("주문번호") # 컬럼 검색 필드 - 주문번호
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합


wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break


################################################



cell_data = google_sheet.acell('H74').value 

driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='도매 매장명']").send_keys(cell_data) # "도매 매장명"에 도매 명 입력
google_sheet.update_acell('I74', 'OK') 
print("74 PASS")
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='도매 매장명']").send_keys(Keys.ENTER) # 검색 적용
google_sheet.update_acell('I75', 'OK') 
print("75 PASS")


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
google_sheet.update_acell('I78', 'OK') 
print("78 PASS")
time.sleep(2)


# 입고 요청 데이터 확인(테이블 - 리스트)

wms_search_list_row_index_ini = 100000
wms_search_list_row = int(0)

wms_search_list_row_index_ini = int(wms_search_list_row_index_ini) # 반복문을 실행하기 위한 int -> 형변환
wms_search_list_row_index_ini = wms_search_list_row_index_ini - 1  # 리스트의 배열이 0부터 시작하기 떄문에 -1

wms_search_list_row_result = driver.find_element(By.CSS_SELECTOR,'div[class="ag-center-cols-container"]')

try:
    while wms_search_list_row_index_ini >= wms_search_list_row : # 리스트 row 수 만큼 실행
        wms_search_list_row = str(wms_search_list_row) # row 0 부터 wms_search_list.find_element에 입력하기 위해 str -> 형변환
        wms_search_list_row_index_str = "div[row-index=\"" + wms_search_list_row + "\"]" # wms_search_list.find_element 에 row 0부터 조회를 위한 값 합치기
        wms_search_list_row_id_result = wms_search_list_row_result.find_element(By.CSS_SELECTOR, wms_search_list_row_index_str) # 리스트 row 마다 ID가져오기
        
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="27"]') # 합계(사입성공수량) : 27
        wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="dueStock"') # 합계(사입성공수량)
        wms_search_list_row_result_check_1 = wms_search_list_row_result_check_1.text
        print("입고 진행현황 - wms_search_list_row_result_check_1   ", wms_search_list_row_result_check_1)
        
        #wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="28"]') # 합계(장끼수량) : 28
        wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="jangkkiStock"]') # 합계(장끼수량)
        wms_search_list_row_result_check_2 = wms_search_list_row_result_check_2.text
        print("입고 진행현황 - wms_search_list_row_result_check_2   ", wms_search_list_row_result_check_2)
        
        
        #wms_search_list_row_result_check_3 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="29"]') # 합계(낱개수량) : 29
        wms_search_list_row_result_check_3 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="12"]') # 합계(낱개수량)
        wms_search_list_row_result_check_3 = wms_search_list_row_result_check_3.text
        print("입고 진행현황 - wms_search_list_row_result_check_3   ", wms_search_list_row_result_check_3)
        
        
        #wms_search_list_row_result_check_4 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="30"]') # 합계(센터입고수량) : 30
        wms_search_list_row_result_check_4 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="inStock"]') # 합계(센터입고수량)
        wms_search_list_row_result_check_4 = wms_search_list_row_result_check_4.text
        print("입고 진행현황 - wms_search_list_row_result_check_4   ", wms_search_list_row_result_check_4)
        
        
        #wms_search_list_row_result_check_5 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="8"]') # 입고상태  : 8
        wms_search_list_row_result_check_5 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="4"]') # 합계(입고상태)
        wms_search_list_row_result_check_5 = wms_search_list_row_result_check_5.text
        print("입고 진행현황 - wms_search_list_row_result_check_5   ", wms_search_list_row_result_check_5)


        print("##########")
        print("입고 진행현황 - 입고 요청 데이터 확인(테이블 - 리스트) 종료")
    

        wms_search_list_row = int(wms_search_list_row) # while을 실행하기 위해 int -> 형변환
        
        if wms_search_list_row == 0 :
            cell_data = google_sheet.acell('H79').value # 도매 상품명_자동화001 -> 합계(사입성공수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I79', 'Pass')
            else:
                google_sheet.update_acell('I79', 'Failed')
    
            
            cell_data = google_sheet.acell('H80').value # 도매 상품명_자동화001 -> 합계(장끼수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I80', 'Pass')
            else:
                google_sheet.update_acell('I80', 'Failed')



            cell_data = google_sheet.acell('H81').value # 도매 상품명_자동화001 -> 합계(낱개수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_3:
                google_sheet.update_acell('I81', 'Pass')
            else:
                google_sheet.update_acell('I81', 'Failed')
    
    
            cell_data = google_sheet.acell('H82').value # 도매 상품명_자동화001 -> 합계(센터입고수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_4:
                google_sheet.update_acell('I82', 'Pass')
            else:
                google_sheet.update_acell('I82', 'Failed')


            cell_data = google_sheet.acell('H83').value # 도매 상품명_자동화001 -> 입고상태 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_5:
                google_sheet.update_acell('I83', 'Pass')
            else:
                google_sheet.update_acell('I83', 'Failed')


    
    
        if wms_search_list_row == 1 :
            cell_data = google_sheet.acell('H85').value # 도매 상품명_자동화002 -> 합계(사입성공수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I85', 'Pass')
            else:
                google_sheet.update_acell('I85', 'Failed')
    
            
            cell_data = google_sheet.acell('H86').value # 도매 상품명_자동화002 -> 합계(장끼수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I86', 'Pass')
            else:
                google_sheet.update_acell('I86', 'Failed')



            cell_data = google_sheet.acell('H87').value # 도매 상품명_자동화002 -> 합계(낱개수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_3:
                google_sheet.update_acell('I87', 'Pass')
            else:
                google_sheet.update_acell('I87', 'Failed')
    
    
            cell_data = google_sheet.acell('H88').value # 도매 상품명_자동화002 -> 합계(센터입고수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_4:
                google_sheet.update_acell('I88', 'Pass')
            else:
                google_sheet.update_acell('I88', 'Failed')


            cell_data = google_sheet.acell('H89').value # 도매 상품명_자동화002 -> 입고상태 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_5:
                google_sheet.update_acell('I89', 'Pass')
            else:
                google_sheet.update_acell('I89', 'Failed')

    
        if wms_search_list_row == 2 :
            cell_data = google_sheet.acell('H91').value # 도매 상품명_자동화003 -> 합계(사입성공수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I91', 'Pass')
            else:
                google_sheet.update_acell('I91', 'Failed')
    
            
            cell_data = google_sheet.acell('H92').value # 도매 상품명_자동화003 -> 합계(장끼수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I92', 'Pass')
            else:
                google_sheet.update_acell('I92', 'Failed')



            cell_data = google_sheet.acell('H93').value # 도매 상품명_자동화003 -> 합계(낱개수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_3:
                google_sheet.update_acell('I93', 'Pass')
            else:
                google_sheet.update_acell('I93', 'Failed')
    
    
            cell_data = google_sheet.acell('H94').value # 도매 상품명_자동화003 -> 합계(센터입고수량)컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_4:
                google_sheet.update_acell('I94', 'Pass')
            else:
                google_sheet.update_acell('I94', 'Failed')


            cell_data = google_sheet.acell('H95').value # 도매 상품명_자동화003 -> 입고상태 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_5:
                google_sheet.update_acell('I95', 'Pass')
            else:
                google_sheet.update_acell('I95', 'Failed')

    
        wms_search_list_row = wms_search_list_row + 1  # row 증가\
            
except:
    pass
   
#리스트의 전체 체크 박스 선택
wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-header-cell.ag-header-cell-sortable.ag-focus-managed') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
wms_test_result.find_element(By.ID, wms_test_result_chk).click()
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합
google_sheet.update_acell('I84', 'OK') 
print("84 PASS")

google_sheet.update_acell('I90', 'OK')
google_sheet.update_acell('I96', 'OK')
# hihida 230102"""


print("##########")
print("(입고 진행현황 - 입고확정가능처리 대상상품만 보기) ] 토클 버튼 실행")
try:
    driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/div[2]/div/div/div[4]/div[2]/label[1]/span[2]').click() # 강제입고확정가능만 보기) ] 토클 버튼 실행
except:
    print("try try try try try try")
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiTypography-root.MuiTypography-body1.MuiFormControlLabel-label')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "입고확정가능처리 대상상품만 보기":
            print("클릭 시도 - 입고확정가능처리 대상상품만 보기)n")
            wms_str_loop.click() # 강제 입고 확장 가능 토글
            print("클릭 완료 - 입고확정가능처리 대상상품만 보기\n")
            break    


time.sleep(5)


print("##########")
print("(입고 진행현황 - 선택상품 입고확정가능 처리) ] 버튼 실행")
try:
    driver.find_element(By.XPATH,'//*[@id="root"]/div/div/div/div/div[2]/div/div/div[4]/div[2]/button[1]').click() # (선택 상품 강제 입고 확정) ] 버튼 종료
    google_sheet.update_acell('I97', 'OK') 
    print("97 PASS")
    time.sleep(3)
    alert = driver.switch_to.alert 
    alert.accept() # 얼럿 확인
    google_sheet.update_acell('I98', 'OK') 
    print("98 PASS")
    time.sleep(3)
    # alert.dismiss()# 취소
except:
    print("try try try try try try")
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeSmall.MuiButton-containedSizeSmall.MuiButtonBase-root')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "선택상품 입고확정가능 처리":
            print("클릭 시도 - 선택상품 입고확정가능 처리\n")
            wms_str_loop.click()
            google_sheet.update_acell('I97', 'OK') 
            print("97 PASS")
            time.sleep(3)
            alert = driver.switch_to.alert 
            alert.accept() # 얼럿 확인
            google_sheet.update_acell('I98', 'OK') 
            print("98 PASS")
            time.sleep(3)
            # alert.dismiss()# 취소
            print("클릭 완료 - 선택상품 입고확정가능 처리\n")
            break
    



try:
    alert = driver.switch_to.alert
    alert.accept() # 얼럿 확인
    time.sleep(3)
except:
    print("try try try try try try")
    pass



########################################################################################################################
##### 입고 관리 - 입고 확정 가능 #####
# 입고 관리 -> 입고 확정 가능 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 확정 가능":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break
        
google_sheet.update_acell('I99', 'OK') 
print("99 PASS")
time.sleep(2)
print("##########")
print("입고 관리 - 입고 확정 가능 이동")

################################################
## 열(컬럼) 메뉴 - 특정 컬럼만 선택
# wms_test_result = driver.find_elements(By.CLASS_NAME, 'ag-side-button-button') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result[0].click() # 중간 오른쪽의 열 컬럼 버튼 클릭

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라짐
# wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-icon.ag-icon-columns') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result.click() # 중간 오른쪽의 열 컬럼 버튼 클릭
# time.sleep(3) # 0[열 컬럼] / 1[필터]


# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라져 다시 수정
# <span class =ag-side-button-labe>에서 열(컬럼), 필터 중 열(컬럼) 찾아 버튼 클릭
time.sleep(2)
wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

time.sleep(3)



driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click()#전체 선택 버튼 다시 클릭해 일단 전체 컬럼 선택 해제
time.sleep(5)


# 체크박스 컬럼 선택
wms_test_result = driver.find_element(By.CSS_SELECTOR, "div>[aria-posinset='1']") # 체크박스 컬럼의 유일값을 찾음, aria-posinset='1'
time.sleep(5)

wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 체크박스 유일값으로 동적으로 변화되는 aria-describedby ID값 취득
time.sleep(5)

wms_test_result.find_element(By.ID, wms_test_result_chk).click()

#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 체크박스 유일값으로 동적으로 변화되는 aria-describedby ID값 취득
#time.sleep(5)

#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() 
time.sleep(5)



# 딜리버드 상품코드 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("딜리버드 상품코드") # 컬럼 검색 필드 - 딜리버드 상품코드 입력
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)


#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 딜리버드 상품코드 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 입고확정가능수량(정상) 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("입고확정가능수량(정상)") # 컬럼 검색 필드 - 입고확정가능수량(정상)
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 입고확정가능수량(정상) 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)



# 주문번호 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("주문번호") # 컬럼 검색 필드 - 주문번호
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합


#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
#time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

################################################
################################################


cell_data = google_sheet.acell('H100').value 

driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='검색어를 입력해주세요.']").send_keys(cell_data) # "도매 매장명"에 도매 명 입력
google_sheet.update_acell('I100', 'OK') 
print("100 PASS")
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='검색어를 입력해주세요.']").send_keys(Keys.ENTER) # 검색 적용
google_sheet.update_acell('I101', 'OK') 
print("101 PASS")
time.sleep(2)

# hihida 230102 WMS 오류로 임시 주석
# driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호

time.sleep(2)

# 입고 요청 데이터 확인(테이블 - 리스트)

wms_search_list_row_index_ini = 100000
wms_search_list_row = int(0)

wms_search_list_row_index_ini = int(wms_search_list_row_index_ini) # 반복문을 실행하기 위한 int -> 형변환
wms_search_list_row_index_ini = wms_search_list_row_index_ini - 1  # 리스트의 배열이 0부터 시작하기 떄문에 -1

wms_search_list_row_result = driver.find_element(By.CSS_SELECTOR,'div[class="ag-center-cols-container"]')

try:
    while wms_search_list_row_index_ini >= wms_search_list_row : # 리스트 row 수 만큼 실행
        wms_search_list_row = str(wms_search_list_row) # row 0 부터 wms_search_list.find_element에 입력하기 위해 str -> 형변환
        wms_search_list_row_index_str = "div[row-index=\"" + wms_search_list_row + "\"]" # wms_search_list.find_element 에 row 0부터 조회를 위한 값 합치기
        wms_search_list_row_id_result = wms_search_list_row_result.find_element(By.CSS_SELECTOR, wms_search_list_row_index_str) # 리스트 row 마다 ID가져오기
        
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="3"]') # 딜리버드 상품코드 : 3
        wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="productExternalProductId"]') # 딜리버드 상품코드
        wms_search_list_row_result_check_1 = wms_search_list_row_result_check_1.text
        print("입고 확정 가능 - wms_search_list_row_result_check_1   ", wms_search_list_row_result_check_1)
        
        #wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="22"]') # 입고확정가능수량(정상) : 22
        wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="normalStock"]') # 입고확정가능수량(정상)
        wms_search_list_row_result_check_2 = wms_search_list_row_result_check_2.text
        print("입고 확정 가능 - wms_search_list_row_result_check_2   ", wms_search_list_row_result_check_2)
        

        print("##########")
        print("입고 확정 가능 - 입고 요청 데이터 확인(테이블 - 리스트) 종료")
    

        wms_search_list_row = int(wms_search_list_row) # while을 실행하기 위해 int -> 형변환
        
        if wms_search_list_row == 0 :
            cell_data = google_sheet.acell('H103').value # 도매 상품명_자동화001 -> 딜리버드 상품코드 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I103', 'Pass')
            else:
                google_sheet.update_acell('I103', 'Failed')
    
            
            cell_data = google_sheet.acell('H104').value # 도매 상품명_자동화001 -> 입고확정가능수량(정상) 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I104', 'Pass')
            else:
                google_sheet.update_acell('I104', 'Failed')

    
    
        if wms_search_list_row == 1 :
            cell_data = google_sheet.acell('H106').value # 도매 상품명_자동화002 -> 딜리버드 상품코드 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I106', 'Pass')
            else:
                google_sheet.update_acell('I106', 'Failed')
    
            
            cell_data = google_sheet.acell('H107').value # 도매 상품명_자동화002 -> 입고확정가능수량(정상) 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I107', 'Pass')
            else:
                google_sheet.update_acell('II107', 'Failed')


    
        if wms_search_list_row == 2 :
            cell_data = google_sheet.acell('H109').value # 도매 상품명_자동화003 -> 딜리버드 상품코드 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I109', 'Pass')
            else:
                google_sheet.update_acell('I109', 'Failed')
    
            
            cell_data = google_sheet.acell('H110').value # 도매 상품명_자동화003 -> 입고확정가능수량(정상) 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I110', 'Pass')
            else:
                google_sheet.update_acell('I110', 'Failed')


    
        wms_search_list_row = wms_search_list_row + 1  # row 증가
        
except:
    pass
   
#리스트의 전체 체크 박스 선택
wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-header-cell.ag-header-cell-sortable.ag-focus-managed') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
wms_test_result_chk = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
time.sleep(5)

wms_test_result.find_element(By.ID, wms_test_result_chk).click()

#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호 체크박스 xPATH값 조합
google_sheet.update_acell('I105', 'OK') 
print("105 PASS")
google_sheet.update_acell('I108', 'OK')
google_sheet.update_acell('I111', 'OK')

# [입고 확정하기] 버튼 클릭
try:
    driver.find_element(By.XPATH, '//*[@id="root"]/div/div/div/div/div[2]/div/div/div[5]/div[2]/button').click()
except:
    print("try try try try try try")
    # 232203 class 값 변경
    # wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeMedium.MuiButton-containedSizeMedium.MuiButtonBase-root')
    wms_test_result = driver.find_elements(By.CSS_SELECTOR,'button.MuiButton-root.MuiButton-contained.MuiButton-containedPrimary.MuiButton-sizeSmall.MuiButton-containedSizeSmall.MuiButtonBase-root')
    for wms_str_loop in wms_test_result:
        #print(wms_str_loop.text,"\n")
        if wms_str_loop.text == "입고 확정하기":
            print("클릭 시도\n")
            wms_str_loop.click()
            print("클릭 완료\n")
            break

google_sheet.update_acell('I112', 'OK') 
print("112 PASS")
time.sleep(3)


alert = driver.switch_to.alert 
alert.accept() # 얼럿 확인
google_sheet.update_acell('I113', 'OK') 
print("113 PASS")
time.sleep(3)
# alert.dismiss()# 취소

try:
    alert = driver.switch_to.alert
    alert.accept() # 얼럿 확인
    time.sleep(3)
except:
    print("try try try try try try")
    pass

print("##########")
print("입고 확정 가능 - (선택 상품 강제 입고 확정) ] 버튼 종료")





########################################################################################################################
##### 입고 관리 - 입고 확정 #####
# 입고 관리 -> 입고 확정 이동
wms_test_result = driver.find_elements(By.CSS_SELECTOR,'span.MuiBox-root')
for wms_str_loop in wms_test_result:
    #print(wms_str_loop.text,"\n")
    if wms_str_loop.text == "입고 확정":
        #print("클릭 시도\n")
        wms_str_loop.click()
        #print("클릭 완료\n")
        break

google_sheet.update_acell('I114', 'OK') 
print("114 PASS")
time.sleep(2)
print("##########")
print("입고 관리 - 입고 확정 이동")

################################################
## 열(컬럼) 메뉴 - 특정 컬럼만 선택
# wms_test_result = driver.find_elements(By.CLASS_NAME, 'ag-side-button-button') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result[0].click() # 중간 오른쪽의 열 컬럼 버튼 클릭

# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라짐
# wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-icon.ag-icon-columns') # 버튼의 상위 그룹의 클래스 네임으로 정보 취득
# wms_test_result.click() # 중간 오른쪽의 열 컬럼 버튼 클릭
# time.sleep(3) # 0[열 컬럼] / 1[필터]


# 230111 WMS 업데이트인 모르지만 ag-icon.ag-icon-columns 명이 사라져 다시 수정
# <span class =ag-side-button-labe>에서 열(컬럼), 필터 중 열(컬럼) 찾아 버튼 클릭
time.sleep(2)
wms_side_columns_button = driver.find_elements(By.CLASS_NAME, 'ag-side-button-label')
#print("wms_side_columns_button", wms_side_columns_button)
for wms_side_button_ck in wms_side_columns_button:
    print("wms_side_button_ck", wms_side_button_ck)
    print("wms_side_button_ck.text",wms_side_button_ck.text)
    if wms_side_button_ck.text == "열(컬럼)":
        print("OK")
        wms_side_button_ck.click()
        break

time.sleep(3)


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='전체 선택']").click()#전체 선택 버튼 다시 클릭해 일단 전체 컬럼 선택 해제
time.sleep(5)


# 딜리버드 상품코드 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("딜리버드 상품코드") # 컬럼 검색 필드 - 딜리버드 상품코드 입력
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)


#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 딜리버드 상품코드 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 입고확정수량(정상) 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("입고확정수량(정상)") # 컬럼 검색 필드 - 입고확정수량(정상)
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 입고확정수량(정상) 체크박스 xPATH값 조합


driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)


# 주문번호 컬럼 선택
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys("주문번호") # 컬럼 검색 필드 - 주문번호 
time.sleep(1)
driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.ENTER)
time.sleep(2)
#time.sleep(5)

#wms_test_result = driver.find_element(By.CLASS_NAME, 'ag-virtual-list-item.ag-column-select-virtual-list-item') # 체크박스의 상위 그룹의 클래스 네임으로 정보 취득
#wms_test_result = wms_test_result.get_attribute("aria-describedby") # 동적으로 변화되는 aria-describedbyD값 취득
#wms_test_result = "//*[@id=\"" + wms_test_result + "\"]""" # 컬럼 선택을 위한 각 체크박스 xPATH값 조합
#driver.find_element(By.XPATH, wms_test_result).click() # 주문번호  체크박스 xPATH값 조합


#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
#time.sleep(2)
#driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='필터 컬럼 입력']").send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
#time.sleep(5)






cell_data = google_sheet.acell('H115').value 

driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='검색어를 입력해주세요.']").send_keys(cell_data) # "도매 매장명"에 도매 명 입력
google_sheet.update_acell('I115', 'OK') 
print("115 PASS")
driver.find_element(By.CSS_SELECTOR, "div>input[placeholder='검색어를 입력해주세요.']").send_keys(Keys.ENTER) # 검색 적용
google_sheet.update_acell('I116', 'OK') 
print("116 PASS")
time.sleep(2)



driver.find_element(By.CSS_SELECTOR, "div>input[aria-label='주문번호 필터 입력']").send_keys(deal_wms_purchase_number) # 테이블(리스트) -> 주문번호
google_sheet.update_acell('I118', 'OK') 
print("118 PASS")
time.sleep(2)


# 입고 확정 쇼핑몰 수 가져오기
time.sleep(10) # 데이터 가져오려면 일정 시간 대기를 해야 로드 할 수 있음

wms_test_result_check = driver.find_elements(By.CLASS_NAME, 'MuiTypography-root.MuiTypography-h2.css-dkmpi')
wms_test_result_check = wms_test_result_check[1].text
print("입고 확정 - wms_test_result_check1", wms_test_result_check,("1"))

wms_test_result_check = wms_test_result_check.replace('입고 확정 쇼핑몰 수 ','')
wms_test_result_check = wms_test_result_check.replace(' 개','') # 텍스트에서 수량만 취득하기 위해 나머지 텍스트 삭제

print("입고 확정 - wms_test_result_check2", wms_test_result_check,("2"))
cell_data = google_sheet.acell('H117').value # 총 도매단가 확인한다.



if cell_data == wms_test_result_check:
    google_sheet.update_acell('I117', 'Pass')
else:
    google_sheet.update_acell('I117', 'Failed')


time.sleep(5)



print("##########")
print("입고 확정 - 입고 확정 쇼핑몰 수 확인 완료")

# hihida 230102"""
# 입고 확정 데이터 확인(테이블 - 리스트)

wms_search_list_row_index_ini = 100000
wms_search_list_row = int(0)

wms_search_list_row_index_ini = int(wms_search_list_row_index_ini) # 반복문을 실행하기 위한 int -> 형변환
wms_search_list_row_index_ini = wms_search_list_row_index_ini - 1  # 리스트의 배열이 0부터 시작하기 떄문에 -1

wms_search_list_row_result = driver.find_element(By.CSS_SELECTOR,'div[class="ag-center-cols-container"]')

try:
    while wms_search_list_row_index_ini >= wms_search_list_row : # 리스트 row 수 만큼 실행
        wms_search_list_row = str(wms_search_list_row) # row 0 부터 wms_search_list.find_element에 입력하기 위해 str -> 형변환
        wms_search_list_row_index_str = "div[row-index=\"" + wms_search_list_row + "\"]" # wms_search_list.find_element 에 row 0부터 조회를 위한 값 합치기
        wms_search_list_row_id_result = wms_search_list_row_result.find_element(By.CSS_SELECTOR, wms_search_list_row_index_str) # 리스트 row 마다 ID가져오기
        
        #wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="2"]') # 딜리버드 상품코드 : 2
        wms_search_list_row_result_check_1 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="productExternalProductId"]') # 딜리버드 상품코드
        wms_search_list_row_result_check_1 = wms_search_list_row_result_check_1.text
        print("입고 확정 - wms_search_list_row_result_check_1   ", wms_search_list_row_result_check_1)
        
        #wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[aria-colindex="21"]') # 입고확정수량(정상) : 21
        wms_search_list_row_result_check_2 = wms_search_list_row_id_result.find_element(By.CSS_SELECTOR, 'div[col-id="normalStock"]') # 입고확정수량(정상)
        wms_search_list_row_result_check_2 = wms_search_list_row_result_check_2.text
        print("입고 확정 - wms_search_list_row_result_check_2   ", wms_search_list_row_result_check_2)
        

        print("##########")
        print("입고 확정 - 입고 확정 데이터 확인(테이블 - 리스트) 종료")
    

        wms_search_list_row = int(wms_search_list_row) # while을 실행하기 위해 int -> 형변환
        
        if wms_search_list_row == 0 :
            cell_data = google_sheet.acell('H119').value # 도매 상품명_자동화001 -> 딜리버드 상품코드 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I119', 'Pass')
            else:
                google_sheet.update_acell('I119', 'Failed')
    
            
            cell_data = google_sheet.acell('H120').value # 도매 상품명_자동화001 -> 입고확정수량(정상) 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I120', 'Pass')
            else:
                google_sheet.update_acell('I120', 'Failed')
    
    
        if wms_search_list_row == 1 :
            cell_data = google_sheet.acell('H121').value # 도매 상품명_자동화002 -> 딜리버드 상품코드 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I121', 'Pass')
            else:
                google_sheet.update_acell('I121', 'Failed')
    
            
            cell_data = google_sheet.acell('H122').value # 도매 상품명_자동화002 -> 입고확정수량(정상) 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I122', 'Pass')
            else:
                google_sheet.update_acell('I122', 'Failed')


    
        if wms_search_list_row == 2 :
            cell_data = google_sheet.acell('H123').value # 도매 상품명_자동화003 -> 딜리버드 상품코드 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_1:
                google_sheet.update_acell('I123', 'Pass')
            else:
                google_sheet.update_acell('I123', 'Failed')
    
            
            cell_data = google_sheet.acell('H124').value # 도매 상품명_자동화003 -> 입고확정수량(정상) 컬럼 확인한다.
            if cell_data == wms_search_list_row_result_check_2:
                google_sheet.update_acell('I124', 'Pass')
            else:
                google_sheet.update_acell('I124', 'Failed')


    
        wms_search_list_row = wms_search_list_row + 1  # row 증가

except:
    pass


#########################################################################################################################
## WMS 이동 -> 딜리버드 이동
#########################################################################################################################
driver.switch_to.window(tabs[0])


#########################################################################################################################
#### 딜리버드 -> 상품 및 재고 #####

try:
    deal_link = driver.find_element(By.LINK_TEXT, ("상품 및 재고"))
    deal_link.click()
    time.sleep(2)
    # driver.find_element(By.XPATH,'//*[@id="navbarSupportedContent"]/ul/li[8]/a').click()
    google_sheet.update_acell('I125', 'OK') 
    print("125 PASS")
    time.sleep(5)
    print("##########")
    print("딜리버드 상품 및 재고 시작")
except:
    print("try try try try try try")
    ##신상마켓 소매 -> 딜리버드 이동
    #신상마켓 로그인
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/header/div/div[2]/div[3]').click() # 로그인 버튼(페이지 상단)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[1]/input').send_keys(deal_seller_login_id) # 모달 / ID 입력
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[2]/input').send_keys(deal_seller_login_password) # 모달 / 비번 입력
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div/div[2]/div[2]/div[2]/div[3]/div[2]/div/button').click() # 모달 / 로그인 버튼
    print("##########")
    print("신상마켓 소매 로그인")

    #딜리버드 바로가기 클릭
    time.sleep(3)
    driver.find_element(By.XPATH, '//*[@id="app"]/div[1]/div[1]/div[1]/div/ul/li[1]/div/span').click()
    print("##########")
    print("딜리버드 이동")
    time.sleep(5)
    
    # driver.find_element(By.XPATH,'//*[@id="navbarSupportedContent"]/ul/li[8]/a').click()
    deal_link = driver.find_element(By.LINK_TEXT, ("상품 및 재고"))
    deal_link.click()
    time.sleep(2)    
    google_sheet.update_acell('I125', 'OK') 
    print("125 PASS")
    time.sleep(5)
    print("##########")
    print("딜리버드상품 및 재고 시작")

# 0126
# 출고 요청 엑셀 파일
deal_wb = openpyxl.load_workbook(deal_out_excel_upload_path)
deal_ws = deal_wb.active

# 출고 요청에게 전달을 위한 변수
deal_sell_product_name_1 = '' # 판매상품명(판매 상품명_자동화상품001)
deal_sell_product_name_2 = '' # 판매상품명(판매 상품명_자동화상품002)
deal_sell_product_name_3 = '' # 판매상품명(판매 상품명_자동화상품003)


# 도매 상품명_자동화001 상품 및 재고 확인
cell_data = google_sheet.acell('H126').value # 구글 시트 -> 딜리버드 상품코드
driver.find_element(By.ID,'search_text').send_keys(cell_data) # 검색어 창에 "딜리버드 상품 코드" 입력
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.ENTER)
google_sheet.update_acell('I126', 'OK') 
print("126 PASS")
time.sleep(2)

# 0126
# 출고 요청 엑셀 파일에 딜리버드 상품 토드 입력
deal_ws['R3'] = cell_data
deal_ws['R4'] = cell_data
deal_product_id_1 = cell_data


deal_table = driver.find_element(By.XPATH, '//*[@id="productList"]') # 리스트(테이블) 전체 경로
deal_tbody = deal_table.find_element(By.TAG_NAME,'tbody')
print("deal_tbody", deal_tbody)
deal_list_count = 0


for tr in deal_tbody.find_elements(By.TAG_NAME,'tr'):
    for td in tr.find_elements(By.TAG_NAME,'td'):
        #print("deal_list_count : ",deal_list_count,"번", td.get_attribute("innerText"),"\n")
        if deal_list_count == 3: # 판매상품명(판매 상품명_자동화상품001)              
            deal_sell_product_name_1 = td.get_attribute("innerText")
        
        if deal_list_count == 8: # 도매 매장명
            cell_data = google_sheet.acell('H127').value
            deal_test_result_check = td.get_attribute("innerText")
            deal_test_result_check = deal_test_result_check.replace('도매 매장 변경','') # 도매 매장명 필드에 [버튼 내용]과 줄바꿈이 있음
            deal_test_result_check = deal_test_result_check.replace("\n",'')
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I127', 'Pass')
            else:
                google_sheet.update_acell('I127', 'Failed')           
            
        if deal_list_count == 13: # 총 재고
            cell_data = google_sheet.acell('H128').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I128', 'Pass')
            else:
                google_sheet.update_acell('I128', 'Failed')           


        if deal_list_count == 14: # 정상재고
            cell_data = google_sheet.acell('H129').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I129', 'Pass')
            else:
                google_sheet.update_acell('I129', 'Failed')           
   
        if deal_list_count == 15: # for문을 더이상 돌지 않도록 설정
            deal_list_count = int(38)

        deal_list_count += 1
        
driver.find_element(By.ID,'search_text').send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
time.sleep(5)



# 도매 상품명_자동화002 상품 및 재고 확인
cell_data = google_sheet.acell('H130').value # 구글 시트 -> 딜리버드 상품코드
driver.find_element(By.ID,'search_text').send_keys(cell_data) # 검색어 창에 "딜리버드 상품 코드" 입력
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.ENTER)
google_sheet.update_acell('I130', 'OK') 
print("130 PASS")
time.sleep(2)

# 0126
# 출고 요청 엑셀 파일에 딜리버드 상품 토드 입력
deal_ws['R5'] = cell_data
deal_ws['R6'] = cell_data
deal_product_id_2 = cell_data


deal_table = driver.find_element(By.XPATH, '//*[@id="productList"]') # 리스트(테이블) 전체 경로
deal_tbody = deal_table.find_element(By.TAG_NAME,'tbody')
deal_list_count = 0


for tr in deal_tbody.find_elements(By.TAG_NAME,'tr'):
    for td in tr.find_elements(By.TAG_NAME,'td'):
        
        if deal_list_count == 3:
            deal_sell_product_name_2 = td.get_attribute("innerText")# 판매상품명(판매 상품명_자동화상품001)
            
        if deal_list_count == 8:
            cell_data = google_sheet.acell('H131').value
            deal_test_result_check = td.get_attribute("innerText")
            deal_test_result_check = deal_test_result_check.replace('도매 매장 변경','') # 도매 매장명 필드에 [버튼 내용]과 줄바꿈이 있음
            deal_test_result_check = deal_test_result_check.replace("\n",'')
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I131', 'Pass')
            else:
                google_sheet.update_acell('I131', 'Failed')           
            
        if deal_list_count == 13:
            cell_data = google_sheet.acell('H132').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I132', 'Pass')
            else:
                google_sheet.update_acell('I132', 'Failed')           


        if deal_list_count == 14:
            cell_data = google_sheet.acell('H133').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I133', 'Pass')
            else:
                google_sheet.update_acell('I133', 'Failed')           
   
        if deal_list_count == 15:
            deal_list_count = int(38)

        deal_list_count += 1
        
driver.find_element(By.ID,'search_text').send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
time.sleep(5)





# 도매 상품명_자동화003 상품 및 재고 확인
cell_data = google_sheet.acell('H134').value # 구글 시트 -> 딜리버드 상품코드
driver.find_element(By.ID,'search_text').send_keys(cell_data) # 검색어 창에 "딜리버드 상품 코드" 입력
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.ENTER)
google_sheet.update_acell('I134', 'OK') 
print("134 PASS")
time.sleep(2)

# 0126
# 출고 요청 엑셀 파일에 딜리버드 상품 토드 입력
deal_ws['R7'] = cell_data
deal_ws['R8'] = cell_data
deal_product_id_3 = cell_data

deal_table = driver.find_element(By.XPATH, '//*[@id="productList"]') # 리스트(테이블) 전체 경로
deal_tbody = deal_table.find_element(By.TAG_NAME,'tbody')
deal_list_count = 0


for tr in deal_tbody.find_elements(By.TAG_NAME,'tr'):
    for td in tr.find_elements(By.TAG_NAME,'td'):
        
        if deal_list_count == 3:
            deal_sell_product_name_3 = td.get_attribute("innerText")# 판매상품명(판매 상품명_자동화상품001)
            
        if deal_list_count == 8:
            cell_data = google_sheet.acell('H135').value
            deal_test_result_check = td.get_attribute("innerText")
            deal_test_result_check = deal_test_result_check.replace('도매 매장 변경','') # 도매 매장명 필드에 [버튼 내용]과 줄바꿈이 있음
            deal_test_result_check = deal_test_result_check.replace("\n",'')
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I135', 'Pass')
            else:
                google_sheet.update_acell('I135', 'Failed')           
            
        if deal_list_count == 13:
            cell_data = google_sheet.acell('H136').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I136', 'Pass')
            else:
                google_sheet.update_acell('I136', 'Failed')           


        if deal_list_count == 14:
            cell_data = google_sheet.acell('H137').value
            deal_test_result_check = td.get_attribute("innerText")
            if cell_data == deal_test_result_check:
                google_sheet.update_acell('I137', 'Pass')
            else:
                google_sheet.update_acell('I137', 'Failed')           
   
        if deal_list_count == 15:
            deal_list_count = int(38)

        deal_list_count += 1
        
driver.find_element(By.ID,'search_text').send_keys(Keys.CONTROL + "a") # 컬럼 검색 필드 - 입력값 전체 선택
time.sleep(2)
driver.find_element(By.ID,'search_text').send_keys(Keys.DELETE) # 컬럼 검색 필드 - 입력값 삭제
time.sleep(5)


# 0126
# 출고 요청 엑셀 파일 저장
deal_out_excel_upload_name_now = datetime.now()


deal_out_excel_upload_name_new = f"자동화_출고요청_QA사입앱자동화_{deal_out_excel_upload_name_now.strftime('%Y%m%d_%H%M%S')}.xlsx"
deal_out_excel_upload_file_path = r"c:\test\\"+deal_out_excel_upload_name_new
deal_wb.save(deal_out_excel_upload_file_path)


# 정상 출고로 보내야 하는 인자값 정리
deal_test_result = driver.find_element(By.XPATH,'//*[@id="page-wrapper"]/div[1]/nav/div/ul/li[1]/strong') # 셀러 이름(대표자)
deal_seller_name = deal_test_result.text




# 출고에 전달할 정보 : 원본 파일
deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정

try:
    info_in_to_out_workbook = openpyxl.load_workbook(deal_in_to_out_excel_path)
except FileNotFoundError:
    # 파일이 없을 경우 새로 생성
    info_in_to_out_workbook = openpyxl.Workbook()


if 'Sheet1' in info_in_to_out_workbook.sheetnames:
    info_in_to_out_sheet = info_in_to_out_workbook.active
else:
    info_in_to_out_sheet = info_in_to_out_workbook.create_sheet('Sheet1')

info_in_to_out_sheet['B1'].value = deal_product_id_1
info_in_to_out_sheet['B2'].value = deal_product_id_2
info_in_to_out_sheet['B3'].value = deal_product_id_3
info_in_to_out_sheet['B4'].value = deal_seller_name
info_in_to_out_sheet['B5'].value = deal_sell_product_name_1
info_in_to_out_sheet['B6'].value = deal_sell_product_name_2
info_in_to_out_sheet['B7'].value = deal_sell_product_name_3

info_in_to_out_workbook.save(deal_in_to_out_excel_path)

print("파일 저장 완료")
print("1. 정상입고 종료")
driver.quit()



# 자동화 정상 입고-사입앱 파일 종료 -> 정상 출고 실행
# 원본
# subprocess.call(['python', '02.AutoTest_OUT_normal.py', json.dumps(file_info)])

# 테스트
# subprocess.call(['python', 'test1.py', json.dumps(file_info)])


#########################
#########################
#########################
#########################
#########################
#########################
#########################
#########################
#########################




#while(True):
#    	pass
 