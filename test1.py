import selenium.webdriver.support.ui as ui

from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.remote.webelement import WebElement
from selenium.webdriver.common.by import By
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.alert import Alert
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog

from openpyxl.styles import Font

import pyautogui
import time
import sys
import json
import re
import openpyxl
import subprocess
import platform


import gspread
from oauth2client.service_account import ServiceAccountCredentials

import requests
from bs4 import BeautifulSoup



#########################################################################################################################
# 사입 요청 파일 다운로드
# https://docs.google.com/spreadsheets/d/19X1duCg7N2npHQHGu_pPcDaji_9pDWdI/edit#gid=1830395100
# 해당 엑셀 파일을 c:\test\ 에 저장



"""
##################### 중요!!!!!
# 테스트 전 복사해서 선언해야 함 deal_test_saip_excel_upload, buyer_wsIdx_name

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화11.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23163' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화12.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23164' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화13.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23165' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화14.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23166' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화15.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23167' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화16.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23168' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화17.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23169' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화18.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23170' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화19.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23171' # 사입 요청한 도매 wsIdx 값

deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값

"""
# deal_test_saip_excel_upload = 'C:\\test\\자동화_사입요청_QA사입앱자동화20.xlsx' # 사입 요청 파일 정보(매핑 : O)
# buyer_wsIdx_name = '23172' # 사입 요청한 도매 wsIdx 값


# 불러오기 창 생성
root = tk.Tk()
root.withdraw()
# excel_upload_file_path = filedialog.askopenfilename()
deal_test_saip_excel_upload = filedialog.askopenfilename()

print("deal_test_saip_excel_upload", deal_test_saip_excel_upload)

# 선택된 파일 열기

# deal_test_saip_excel_upload = openpyxl.load_workbook(excel_upload_file_path)
# org_workbook = openpyxl.load_workbook(excel_upload_file_path)
# org_worsheet = org_workbook.active

# 사입 요청한 도매 wsIdx 값
#buyer_wsIdx_name = input("Enter 도매 wsIndex number: ")

test_os = platform.system()

if test_os == "Windows":
    print("윈도우입니다.", test_os)
elif test_os == "Darwin":
    print("맥입니다.",test_os)
else:
    print("윈도우나 맥이 아닙니다.",test_os)


'''
#########################################################################################################################
# 테스트 기본 설정
#########################################################################################################################
# 자동화 실행 시 기본 정보 로드 : ex)계정 정보 입력, 구글 접속 정보
deal_admin_login_id = ''
deal_admin_login_password = ''
deal_admin_url = 'https://dealibird.qa.sinsang.market/ssm_admins/sign_in'
deal_seller_login_id = ''
deal_seller_login_password = ''
deal_seller_url = 'https://vat.qa.sinsang.market/'

# WMS 테스트 기본 설정
wms_login_id = '' 									# WMS 로그인 ID
wms_login_passWord = ''  						# WMS 로그인 비번
wms_url = 'https://matrix-web.qa.sinsang.market/signin'

info_file_path = 'C:\\test\\info.xlsx'
try:
    info_workbook = openpyxl.load_workbook(info_file_path)
except FileNotFoundError:
    # 파일이 없을 경우 새로 생성
    info_workbook = openpyxl.Workbook()


if 'Sheet1' in info_workbook.sheetnames:
    info_sheet = info_workbook.active
else:
    info_sheet = info_workbook.create_sheet('Sheet1')


# info_file = askopenfilename(title="Select an Excel file", filetypes=[("Excel files", "C:\\test\\info.xlsx")])

# info_file = askopenfilename(info_file_path)

# 엑셀 파일 열기
# info_workbook = openpyxl.load_workbook(info_file)

# 로그인 정보 입력 받기
#deal_admin_login_id = info_workbook.active['A1']
#deal_admin_login_password = info_workbook.active['B1']
deal_admin_login_id_cell = info_sheet['A1']
deal_admin_login_password_cell = info_sheet['A2']

deal_seller_login_id_cell = info_sheet['A3']
deal_seller_login_password_cell = info_sheet['A4']

wms_login_id_cell = info_sheet['A5']
wms_login_passWord_cell = info_sheet['A6']

font_color = Font(color='FFFFFF')
# print("0번째", deal_admin_login_id_cell)
# print("0번째", deal_admin_login_password_cell)

##### 어드민 계정 정보 #####
if deal_admin_login_id_cell.value is None:
    deal_admin_login_id_input = (input("어드민 로그인 ID: "))
    deal_admin_login_id_cell.value = deal_admin_login_id_input
    info_sheet['A1'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_id = deal_admin_login_id_input
    # print("1번째", deal_admin_login_id)
else:
    deal_admin_login_id = deal_admin_login_id_cell.value
    #print("2번째", deal_admin_login_id)


if deal_admin_login_password_cell.value is None:
    deal_admin_login_password_input = (input("어드민 로그인 비밀번호: "))
    deal_admin_login_password_cell.value = deal_admin_login_password_input
    info_sheet['A2'].font = font_color
    info_workbook.save(info_file_path)
    deal_admin_login_password = deal_admin_login_password_input
    #print("1번째", deal_admin_login_password)

else:
    deal_admin_login_password = deal_admin_login_password_cell.value
    #print("2번째", deal_admin_login_password)

##### 셀러 계정 정보 #####
if deal_seller_login_id_cell.value is None:
    deal_seller_login_id_input = (input("셀러 로그인 ID: "))
    deal_seller_login_id_cell.value = deal_seller_login_id_input
    info_sheet['A3'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_id = deal_seller_login_id_input
    # print("1번째", deal_seller_login_id)
else:
    deal_seller_login_id = deal_seller_login_id_cell.value
    #print("2번째", deal_seller_login_id)


if deal_seller_login_password_cell.value is None:
    deal_seller_login_password_input = (input("셀러 로그인 비밀번호: "))
    deal_seller_login_password_cell.value = deal_seller_login_password_input
    info_sheet['A4'].font = font_color
    info_workbook.save(info_file_path)
    deal_seller_login_password = deal_seller_login_password_input
    #print("1번째", deal_seller_login_password)

else:
    deal_seller_login_password = deal_seller_login_password_cell.value
    #print("2번째", deal_seller_login_password)


##### WMS 계정 정보 #####
if wms_login_id_cell.value is None:
    wms_login_id_input = (input("WMS 로그인 ID: "))
    wms_login_id_cell.value = wms_login_id_input
    info_sheet['A5'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_id = wms_login_id_input
    # print("1번째", wms_login_id)
else:
    wms_login_id = wms_login_id_cell.value
    #print("2번째", wms_login_id)


if wms_login_passWord_cell.value is None:
    wms_login_passWord_input = (input("WMS 로그인 비밀번호: "))
    wms_login_passWord_cell.value = wms_login_passWord_input
    info_sheet['A6'].font = font_color
    info_workbook.save(info_file_path)
    wms_login_passWord = wms_login_passWord_input
    #print("1번째", wms_login_passWord)

else:
    wms_login_passWord = wms_login_passWord_cell.value
    #print("2번째", wms_login_passWord)




# 출고 관련 정보(배송 요청) : 원본 파일
deal_out_excel_upload_path = r"c:\test\auto_out_org_file.xlsx" # Excel 파일 경로 설정

# 출고에 전달할 정보 : 원본 파일
deal_in_to_out_excel_path = r"c:\test\info_in_to_out.xlsx" # Excel 파일 경로 설정



#########################################################################################################################
# 사입 요청 엑셀 파일 -> 사입앱 -> 사입 수량, 도매 단가, 최종 금액 변경에 대한 정보 입력

buyer_workbook = openpyxl.load_workbook(deal_test_saip_excel_upload)
buyer_worsheet = buyer_workbook.active
'''
while(True):
    pass
 